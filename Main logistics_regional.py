from pathlib import Path
import os
import json
import requests
import time
import sys
from datetime import datetime
import threading
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
from functools import lru_cache
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from geopy.distance import geodesic

def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)

# =============================================================================
# Regional Definitions for Equipment Optimization
# =============================================================================

REGIONS = {
    'Region 1 (East/South)': {
        'warehouses': ['CHICAGO', 'BOSTON', 'PHILADELPH', 'NEWYORK', 'ORLANDO', 'FTLAUDER', 'WDC', 'NASHVILLE'],
        'description': 'Eastern and Southern US warehouses'
    },
    'Region 2 (West/Central)': {
        'warehouses': ['ANAHEIM', 'SANFRAN', 'PHOENIX', 'LASVEGAS', 'DALLAS'],
        'description': 'Western and Central US warehouses'
    }
}

# =============================================================================
# OpenRouteService Implementation for Road-Based Distances
# =============================================================================

class OpenRouteService:
    """
    OpenRouteService API for road-based distance calculations
    Sign up for free API key at: https://openrouteservice.org/dev/#/signup
    Free tier: 2000 requests per day
    """
    
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://api.openrouteservice.org/v2"
        self.session = requests.Session()
        self.session.headers.update({
            'Authorization': api_key,
            'Content-Type': 'application/json'
        })
        print(f"✅ OpenRouteService initialized with API key: {api_key[:8]}...")
    
    @lru_cache(maxsize=500)  # Cache results to avoid repeated API calls
    def get_driving_distance(self, start_coords: Tuple[float, float], 
                           end_coords: Tuple[float, float]) -> float:
        """
        Get driving distance between two points in miles
        
        Args:
            start_coords: (latitude, longitude) 
            end_coords: (latitude, longitude)
            
        Returns:
            Distance in miles, or fallback to geodesic if API fails
        """
        try:
            # ORS uses longitude, latitude format (opposite of lat/lon)
            start_lon, start_lat = start_coords[1], start_coords[0]
            end_lon, end_lat = end_coords[1], end_coords[0]
            
            url = f"{self.base_url}/directions/driving-car"
            
            body = {
                "coordinates": [[start_lon, start_lat], [end_lon, end_lat]],
                "format": "json",
                "units": "mi"  # Return in miles
            }
            
            # Add small delay to be respectful to the API
            time.sleep(0.1)
            
            response = self.session.post(url, json=body, timeout=15)
            response.raise_for_status()
            
            data = response.json()
            
            # Extract distance from response
            if 'routes' in data and len(data['routes']) > 0:
                distance_miles = data['routes'][0]['summary']['distance']
                print(f"🛣️  ORS: {start_coords} → {end_coords} = {distance_miles:.1f} miles")
                return distance_miles
            else:
                print(f"⚠️  No routes found in ORS response for {start_coords} → {end_coords}")
                return self._fallback_distance(start_coords, end_coords)
                
        except requests.exceptions.RequestException as e:
            print(f"⚠️  ORS API error: {e}")
            return self._fallback_distance(start_coords, end_coords)
        except Exception as e:
            print(f"⚠️  ORS processing error: {e}")
            return self._fallback_distance(start_coords, end_coords)
    
    def _fallback_distance(self, coord1: Tuple[float, float], coord2: Tuple[float, float]) -> float:
        """Fallback to geodesic distance if API fails"""
        try:
            distance = geodesic(coord1, coord2).miles
            print(f"📏 Fallback to geodesic: {coord1} → {coord2} = {distance:.1f} miles")
            return distance
        except Exception as e:
            print(f"❌ Error calculating fallback distance: {e}")
            return 999999.0
    
    def test_connection(self) -> bool:
        """Test if the API key and connection work"""
        try:
            # Test with a simple route (New York to Philadelphia)
            test_distance = self.get_driving_distance((40.7128, -74.0060), (39.9526, -75.1652))
            if test_distance < 999999:
                print(f"✅ ORS connection test successful! NY→Philly = {test_distance:.1f} miles")
                return True
            else:
                print("❌ ORS connection test failed")
                return False
        except Exception as e:
            print(f"❌ ORS connection test error: {e}")
            return False

# Set your OpenRouteService API key here
# Get free API key at: https://openrouteservice.org/dev/#/signup
ORS_API_KEY = "eyJvcmciOiI1YjNjZTM1OTc4NTExMTAwMDFjZjYyNDgiLCJpZCI6IjBkMDU3OGYxOTMzYjQ5ZmQ5NzdkMWU0OTUyMGI2NGY1IiwiaCI6Im11cm11cjY0In0="  # REPLACE WITH YOUR ACTUAL API KEY

# Configuration
WAREHOUSE_COORDS = {
    'ANAHEIM': (33.8366, -117.9143),
    'BOSTON': (42.3601, -71.0589),
    'CHICAGO': (41.8781, -87.6298),
    'DALLAS': (32.7767, -96.7970),
    'FTLAUDER': (26.1224, -80.1373),
    'LASVEGAS': (36.1699, -115.1398),
    'NASHVILLE': (36.1627, -86.7816),
    'NEWYORK': (40.7128, -74.0060),
    'ORLANDO': (28.5383, -81.3792),
    'PHOENIX': (33.4484, -112.0740),
    'PHILADELPH': (39.9526, -75.1652),
    'SANFRAN': (37.7749, -122.4194),
    'WDC': (38.9072, -77.0369)
}

NON_DATE_FIELDS = {
    'Equipment', 'Location', 'Description', 'Total\nOwn', 'Total Own', 'Out', 'On\nHand', 'On Hand',
    'Avail\nToday', 'To Ship\nToday', 'To Ship Today', 'Committed', 'Late\nReturn', 'Late Return',
    'QC', 'To Return\nToday', 'To Return Today', 'Department', 'Category', 'Owner', 'Sub Out',
    'Repair', 'Inspection', 'Quarantined', 'Reserved Barcode', 'In Truck', 'Locked Attach.',
    'Late Ship', 'Late Sub Ship', 'Late Sub Return', 'Late Repair', 'Late Sub Repair',
    '%1', '%2', '%3', '%4', '%5', '%6', '%7', '%8', '%9', '%',
    'Qty.1', 'Qty.2', 'Qty.3', 'Qty.4', 'Qty.5', 'Qty.6', 'Qty.7', 'Qty.8', 'Qty.9',
    'Qty1', 'Qty2', 'Qty3', 'Qty4', 'Qty5', 'Qty6', 'Qty7', 'Qty8', 'Qty9'
}

CONFIG_FILE = resource_path("logistics_config.json")
WEIGHTS_FILE = resource_path("equipment_weights.json")



@dataclass
class PullResult:
    warehouse: str
    quantity: int
    distance: int
    available_quantity: int
    description: str = ""
    is_secondary: bool = False
    weight_per_unit: float = 0.0
    total_weight: float = 0.0
    region: str = ""  # Region this warehouse belongs to
    min_avail_qty: int = 0  # Min Avail Qty from Excel (indicates partial availability)
    raw_available: int = 0  # Raw Qty Available from Excel
    is_partial_availability: bool = False  # True if Min Avail Qty > Qty Available (needs backfill)
    is_backfill: bool = False  # True if this pull is a backfill for partial availability
    backfill_for_warehouse: str = ""  # If is_backfill=True, which warehouse is being backfilled
    
    # New fields for daily availability analysis
    daily_timeline: Dict[str, int] = field(default_factory=dict)
    shortage_dates: List[str] = field(default_factory=list)
    shortage_ranges: List[Dict] = field(default_factory=list)  # [{'start': str, 'end': str, 'qty': int}]
    daily_shortages: Dict[str, int] = field(default_factory=dict)
    max_daily_shortage: int = 0


@dataclass
class OptimizationScenario:
    """Container for a complete optimization scenario"""
    name: str
    description: str
    results: List['OptimizationResult']
    total_distance: float
    fill_rate: float
    total_trips: int
    consolidated_trips: int
    total_weight: float


@dataclass
class OptimizationResult:
    equipment_code: str
    equipment_description: str
    requested_quantity: int
    pulls: List[PullResult]
    shortfall: int
    total_distance: float
    weight_per_unit: float = 0.0
    total_weight: float = 0.0


class WeightManager:
    """Handle equipment weight data"""
    
    def __init__(self):
        self.weights = {}
        self.load_weights()
    
    def load_weights(self):
        """Load equipment weights from JSON file"""
        try:
            if os.path.exists(WEIGHTS_FILE):
                with open(WEIGHTS_FILE, 'r') as f:
                    data = json.load(f)
                    self.weights = data.get('equipment_weights', {})
                print(f"Loaded {len(self.weights)} equipment weights from {WEIGHTS_FILE}")
            else:
                print(f"No weights file found at {WEIGHTS_FILE}, starting with empty weights")
        except Exception as e:
            print(f"Error loading weights: {e}")
            self.weights = {}
    
    def save_weights(self):
        """Save equipment weights to JSON file"""
        try:
            data = {
                'equipment_weights': self.weights,
                'last_updated': datetime.now().isoformat()
            }
            with open(WEIGHTS_FILE, 'w') as f:
                json.dump(data, f, indent=2)
            print(f"Saved {len(self.weights)} equipment weights to {WEIGHTS_FILE}")
        except Exception as e:
            print(f"Error saving weights: {e}")
    
    def get_weight(self, equipment_code: str) -> float:
        """Get weight for equipment code"""
        return self.weights.get(equipment_code, 0.0)
    
    def set_weight(self, equipment_code: str, weight: float):
        """Set weight for equipment code"""
        self.weights[equipment_code] = weight
        self.save_weights()
    
    def update_weights(self, weight_dict: Dict[str, float]):
        """Update multiple weights at once"""
        self.weights.update(weight_dict)
        self.save_weights()


class LogisticsEngine:
    """Enhanced business logic with OpenRouteService road-based distances and regional optimization"""

    def __init__(self):
        self.weight_manager = WeightManager()
        
        # Initialize OpenRouteService for road-based distances
        try:
            self.ors = OpenRouteService(ORS_API_KEY)
            # Test the connection
            if self.ors.test_connection():
                print("🛣️  Road-based distance calculation enabled via OpenRouteService")
                self.use_road_distances = True
            else:
                print("⚠️  ORS connection failed, falling back to geodesic distances")
                self.use_road_distances = False
        except Exception as e:
            print(f"⚠️  Failed to initialize OpenRouteService: {e}")
            print("📏 Using geodesic distances (straight line)")
            self.ors = None
            self.use_road_distances = False

    def get_warehouse_region(self, warehouse: str) -> str:
        """Get the region a warehouse belongs to"""
        for region_name, region_data in REGIONS.items():
            if warehouse in region_data['warehouses']:
                return region_name
        return "Unknown Region"

    def calculate_distance(self, coord1: Tuple[float, float], coord2: Tuple[float, float]) -> float:
        """Calculate distance between two coordinates using road routing if available"""
        if self.use_road_distances and self.ors:
            try:
                return self.ors.get_driving_distance(coord1, coord2)
            except Exception as e:
                print(f"⚠️  Road distance calculation failed, using geodesic: {e}")
        
        # Fallback to geodesic distance
        try:
            distance = geodesic(coord1, coord2).miles
            return distance
        except Exception as e:
            print(f"❌ Error calculating distance: {e}")
            return 999999.0
    
    def verify_warehouse_distances(self, destination: str) -> Dict[str, float]:
        """Verify and return all warehouse distances from destination"""
        if destination not in WAREHOUSE_COORDS:
            raise ValueError(f"Destination '{destination}' not found in warehouse coordinates")
        
        dest_coords = WAREHOUSE_COORDS[destination]
        distances = {}
        
        distance_type = "ROAD" if self.use_road_distances else "GEODESIC"
        print(f"\n=== {distance_type} DISTANCE VERIFICATION FOR {destination} ===")
        print(f"Destination coordinates: {dest_coords}")
        if self.use_road_distances:
            print("🛣️  Using OpenRouteService for road-based distances")
            print("⏱️  This may take a moment for API calls...")
        else:
            print("📏 Using geodesic distances (straight line)")
        print("-" * 50)
        
        total_warehouses = len(WAREHOUSE_COORDS)
        for i, (warehouse, coords) in enumerate(WAREHOUSE_COORDS.items(), 1):
            if warehouse == destination:
                distances[warehouse] = 0.0
                print(f"{warehouse:<12}: {0:>8.1f} miles (LOCAL) [{i}/{total_warehouses}]")
            else:
                try:
                    if self.use_road_distances:
                        print(f"🛣️  Calculating road distance to {warehouse}... [{i}/{total_warehouses}]", end="", flush=True)
                    
                    distance = self.calculate_distance(coords, dest_coords)
                    distances[warehouse] = distance
                    
                    if distance >= 900000:
                        print(f"\r{warehouse:<12}: ERROR - Distance calculation failed [{i}/{total_warehouses}]")
                    else:
                        distance_type_short = "ROAD" if self.use_road_distances else "GEO"
                        print(f"\r{warehouse:<12}: {distance:>8.1f} miles ({distance_type_short}) [{i}/{total_warehouses}]")
                        
                except Exception as e:
                    print(f"\r{warehouse:<12}: ERROR - {e} [{i}/{total_warehouses}]")
                    distances[warehouse] = 999999.0
        
        print("-" * 50)
        
        sorted_warehouses = sorted(distances.items(), key=lambda x: x[1])
        print("WAREHOUSES BY DISTANCE:")
        for i, (wh, dist) in enumerate(sorted_warehouses, 1):
            if dist == 0:
                print(f"{i:>2}. {wh:<12}: LOCAL")
            elif dist > 900000:
                print(f"{i:>2}. {wh:<12}: ERROR")
            else:
                print(f"{i:>2}. {wh:<12}: {dist:>8.1f} miles")
        
        print("=" * 50)
        return distances
    
    def get_equipment_description(self, df: pd.DataFrame, equipment: str) -> str:
        """Get description for equipment from the dataframe"""
        try:
            filtered_df = df[df['Equipment'] == equipment]
            if not filtered_df.empty and 'Description' in df.columns:
                description = filtered_df['Description'].iloc[0]
                return str(description) if pd.notnull(description) else ""
            return ""
        except:
            return ""
    
    def get_min_avail_qty(self, df: pd.DataFrame, equipment: str, location: str) -> int:
        """Get minimum available quantity threshold for equipment at specific location"""
        filtered_df = df[
            (df['Equipment'] == equipment) &
            (df['Location'].str.upper() == location.upper())
        ].copy()

        if filtered_df.empty:
            location_variations = {
                'NEWYORK': ['NEW YORK', 'NY', 'NYC'],
                'SANFRAN': ['SAN FRANCISCO', 'SF', 'SAN FRAN'],
                'LASVEGAS': ['LAS VEGAS', 'VEGAS'],
                'FTLAUDER': ['FT LAUDERDALE', 'FORT LAUDERDALE', 'FT LAUDER'],
                'PHILADELPH': ['PHILADELPHIA', 'PHILLY']
            }

            location_upper = location.upper()
            if location_upper in location_variations:
                for variation in location_variations[location_upper]:
                    filtered_df = df[
                        (df['Equipment'] == equipment) &
                        (df['Location'].str.upper() == variation)
                    ].copy()
                    if not filtered_df.empty:
                        break

        if filtered_df.empty:
            return 0

        # Check if Min Avail Qty column exists
        min_avail_col = None
        for col in df.columns:
            if str(col).strip().lower() in ['min avail qty', 'min_avail_qty', 'minavailqty']:
                min_avail_col = col
                break

        if min_avail_col is None:
            return 0

        try:
            min_qty = filtered_df[min_avail_col].iloc[0]
            if pd.notnull(min_qty):
                return int(min_qty)
            return 0
        except Exception as e:
            return 0

    def get_available_quantity(self, df: pd.DataFrame, equipment: str,
                             location: str, date_columns: List[str]) -> int:
        """Get available quantity for equipment at specific location based on date range

        Returns the Qty Available value from the Excel file.
        Use get_min_avail_qty() separately to check for partial availability scenarios.
        """
        filtered_df = df[
            (df['Equipment'] == equipment) &
            (df['Location'].str.upper() == location.upper())
        ].copy()

        if filtered_df.empty:
            location_variations = {
                'NEWYORK': ['NEW YORK', 'NY', 'NYC'],
                'SANFRAN': ['SAN FRANCISCO', 'SF', 'SAN FRAN'],
                'LASVEGAS': ['LAS VEGAS', 'VEGAS'],
                'FTLAUDER': ['FT LAUDERDALE', 'FORT LAUDERDALE', 'FT LAUDER'],
                'PHILADELPH': ['PHILADELPHIA', 'PHILLY']
            }

            location_upper = location.upper()
            if location_upper in location_variations:
                for variation in location_variations[location_upper]:
                    filtered_df = df[
                        (df['Equipment'] == equipment) &
                        (df['Location'].str.upper() == variation)
                    ].copy()
                    if not filtered_df.empty:
                        break

        if filtered_df.empty:
            return 0

        try:
            numeric_data = filtered_df[date_columns].apply(pd.to_numeric, errors='coerce')

            if len(date_columns) == 1:
                result = numeric_data.iloc[0, 0]
                return int(result) if pd.notnull(result) and result >= 0 else 0
            else:
                date_values = numeric_data.iloc[0].values
                valid_values = [val for val in date_values if pd.notnull(val) and val >= 0]

                if not valid_values:
                    return 0

                if len(valid_values) == len(date_columns):
                    result = min(valid_values)
                else:
                    print(f"  WARNING: {equipment} at {location} has gaps in availability "
                          f"({len(valid_values)}/{len(date_columns)} dates have data)")
                    return 0

                return int(result) if result >= 0 else 0

        except Exception as e:
            print(f"Error processing availability for {equipment} at {location}: {e}")
            return 0

    def analyze_daily_availability(self, df: pd.DataFrame, equipment: str,
                                 location: str, date_columns: List[str],
                                 needed_qty: int) -> Dict:
        """
        Analyze availability day-by-day to find specific gaps and shortages.
        Returns a dictionary with detailed timeline and shortage info.
        """
        result = {
            'timeline': {},
            'shortage_dates': [],
            'daily_shortages': {},
            'max_shortage': 0
        }
        
        if not date_columns:
            return result
            
        filtered_df = df[
            (df['Equipment'] == equipment) &
            (df['Location'].str.upper() == location.upper())
        ].copy()
        
        # Handle location variations if empty
        if filtered_df.empty:
            location_variations = {
                'NEWYORK': ['NEW YORK', 'NY', 'NYC'],
                'SANFRAN': ['SAN FRANCISCO', 'SF', 'SAN FRAN'],
                'LASVEGAS': ['LAS VEGAS', 'VEGAS'],
                'FTLAUDER': ['FT LAUDERDALE', 'FORT LAUDERDALE', 'FT LAUDER'],
                'PHILADELPH': ['PHILADELPHIA', 'PHILLY']
            }
            location_upper = location.upper()
            if location_upper in location_variations:
                for variation in location_variations[location_upper]:
                    filtered_df = df[
                        (df['Equipment'] == equipment) &
                        (df['Location'].str.upper() == variation)
                    ].copy()
                    if not filtered_df.empty:
                        break
        
        if filtered_df.empty:
            return result
            
        try:
            # Extract daily values
            numeric_data = filtered_df[date_columns].apply(pd.to_numeric, errors='coerce')
            daily_values = numeric_data.iloc[0].to_dict()
            
            result['timeline'] = daily_values
            
            # Analyze for shortages against needed_qty
            max_shortage = 0
            for date_col, avail_qty in daily_values.items():
                if pd.isna(avail_qty):
                    avail_qty = 0
                
                # Check if available quantity is less than what we need
                if avail_qty < needed_qty:
                    shortage = needed_qty - avail_qty
                    result['shortage_dates'].append(date_col)
                    result['daily_shortages'][date_col] = shortage
                    if shortage > max_shortage:
                        max_shortage = shortage
            
            result['max_shortage'] = max_shortage
            
            # --- Consolidate Shortage Ranges ---
            # Group consecutive shortage dates into ranges
            ranges = []
            if result['shortage_dates']:
                current_range = None
                
                # Iterate through ALL date columns in order to preserve sequence
                for date_col in date_columns:
                    if date_col in result['daily_shortages']:
                        qty = result['daily_shortages'][date_col]
                        
                        if current_range is None:
                            # Start new range
                            current_range = {
                                'start': date_col,
                                'end': date_col,
                                'dates': [date_col],
                                'max_qty': qty
                            }
                        else:
                            # Extend current range
                            current_range['end'] = date_col
                            current_range['dates'].append(date_col)
                            current_range['max_qty'] = max(current_range['max_qty'], qty)
                    else:
                        # End current range if it exists
                        if current_range is not None:
                            ranges.append(current_range)
                            current_range = None
                
                # Append final range if active
                if current_range is not None:
                    ranges.append(current_range)
            
            result['shortage_ranges'] = ranges
            
        except Exception as e:
            print(f"Error analyzing daily availability for {equipment} at {location}: {e}")
            
        return result
    
    def find_warehouse_options_for_equipment(self, df: pd.DataFrame, equipment: str, 
                                           destination: str, date_columns: List[str],
                                           needed_qty: int, region_filter: str = None) -> List[Tuple[str, int, float, float]]:
        """Find all warehouse options for a specific equipment, sorted by score (includes destination)"""
        options = []
        dest_coords = WAREHOUSE_COORDS[destination]
        
        # Get warehouses to consider based on region filter
        warehouses_to_check = WAREHOUSE_COORDS.keys()
        if region_filter and region_filter in REGIONS:
            region_warehouses = REGIONS[region_filter]['warehouses']
            # Always include destination even if not in selected region
            if destination not in region_warehouses:
                warehouses_to_check = region_warehouses + [destination]
            else:
                warehouses_to_check = region_warehouses
            print(f"  🌍 Regional filter: {region_filter} - checking warehouses: {warehouses_to_check}")
        
        for warehouse in warehouses_to_check:
            if warehouse not in WAREHOUSE_COORDS:
                continue
                
            coords = WAREHOUSE_COORDS[warehouse]
            available = self.get_available_quantity(df, equipment, warehouse, date_columns)
            if available > 0:
                if warehouse == destination:
                    distance = 0.0
                else:
                    distance = self.calculate_distance(coords, dest_coords)
                    if distance >= 900000:  # Skip invalid distances
                        continue
                
                # Score based on availability and distance (same as optimization logic)
                # Higher availability and lower distance = better score
                fulfillment_ratio = min(available / needed_qty, 1.0)
                score = fulfillment_ratio / (distance + 1) * 1000  # Scale for readability
                options.append((warehouse, available, distance, score))
        
        # Sort by score (higher is better) - this gives us the true ranking
        options.sort(key=lambda x: x[3], reverse=True)
        return [(w, avail, dist, score) for w, avail, dist, score in options]
    
    def create_pull_result(self, warehouse: str, pull_qty: int, distance: int,
                          total_available: int, equipment_description: str,
                          equipment_code: str, df: pd.DataFrame = None,
                          date_columns: List[str] = None,
                          is_backfill: bool = False,
                          backfill_for_warehouse: str = "") -> PullResult:
        """Create a PullResult with weight, region, and partial availability information"""
        weight_per_unit = self.weight_manager.get_weight(equipment_code)
        total_weight = pull_qty * weight_per_unit
        region = self.get_warehouse_region(warehouse)

        # Get min avail qty and check for partial availability
        min_avail_qty = 0
        raw_available = total_available
        is_partial_availability = False

        if df is not None and date_columns is not None:
            min_avail_qty = self.get_min_avail_qty(df, equipment_code, warehouse)
            raw_available = self.get_available_quantity(df, equipment_code, warehouse, date_columns)
            # Partial availability: Qty Available > 0 AND Min Avail Qty < 0 (negative = backfill needed)
            if raw_available > 0 and min_avail_qty < 0:
                is_partial_availability = True
            
            # Perform daily availability analysis
            daily_analysis = self.analyze_daily_availability(df, equipment_code, warehouse, date_columns, pull_qty)

        return PullResult(
            warehouse=warehouse,
            quantity=pull_qty,
            distance=distance,
            available_quantity=total_available,
            description=equipment_description,
            weight_per_unit=weight_per_unit,
            total_weight=total_weight,
            region=region,
            min_avail_qty=min_avail_qty,
            raw_available=raw_available,
            is_partial_availability=is_partial_availability,
            is_backfill=is_backfill,
            backfill_for_warehouse=backfill_for_warehouse,
            daily_timeline=daily_analysis.get('timeline', {}) if daily_analysis else {},
            shortage_dates=daily_analysis.get('shortage_dates', []) if daily_analysis else [],
            shortage_ranges=daily_analysis.get('shortage_ranges', []) if daily_analysis else [],
            daily_shortages=daily_analysis.get('daily_shortages', {}) if daily_analysis else {},
            max_daily_shortage=daily_analysis.get('max_shortage', 0) if daily_analysis else 0
        )

    def handle_partial_availability(self, df: pd.DataFrame, equipment_code: str,
                                    equipment_description: str, warehouse: str,
                                    destination: str, date_columns: List[str],
                                    warehouse_distances: Dict[str, float],
                                    pulls: List[PullResult],
                                    region_filter: str = None) -> int:
        """
        Handle partial availability scenario.

        Logic: If Qty Available is POSITIVE and Min Avail Qty is NEGATIVE:
        - Qty Available = units the location can supply for full rental period
        - Min Avail Qty (negative) = units that need to be backfilled

        Example: Qty Available = 1, Min Avail Qty = -4
        - Orlando can supply 1 unit for full period
        - 4 units need to be backfilled from another location

        Returns the backfill quantity needed, and adds backfill pull to the pulls list.
        """
        qty_available = self.get_available_quantity(df, equipment_code, warehouse, date_columns)
        min_avail_qty = self.get_min_avail_qty(df, equipment_code, warehouse)

        # Check if this is a partial availability scenario
        # Qty Available > 0 (positive) AND Min Avail Qty < 0 (negative) = backfill needed
        if qty_available > 0 and min_avail_qty < 0:
            backfill_needed = abs(min_avail_qty)  # Convert negative to positive

            print(f"  🔄 PARTIAL AVAILABILITY DETECTED:")
            print(f"     {warehouse}: Can supply {qty_available} units for full period")
            print(f"     Min Avail Qty: {min_avail_qty} (backfill indicator)")
            print(f"     Backfill needed: {backfill_needed} units from another location")

            # Find closest warehouse to provide backfill (excluding the current warehouse)
            backfill_options = []
            for wh, dist in warehouse_distances.items():
                if wh == warehouse:
                    continue

                # Check region filter if applicable
                if region_filter:
                    wh_region = self.get_warehouse_region(wh)
                    if wh_region != region_filter and wh != destination:
                        continue

                wh_available = self.get_available_quantity(df, equipment_code, wh, date_columns)
                if wh_available > 0:
                    backfill_options.append((wh, wh_available, dist))

            # Sort by distance (closest first)
            backfill_options.sort(key=lambda x: x[2])

            if backfill_options:
                backfill_warehouse, backfill_avail, backfill_dist = backfill_options[0]
                backfill_qty = min(backfill_needed, backfill_avail)

                print(f"     ✓ Backfill source: {backfill_warehouse} ({backfill_qty} units, {int(backfill_dist)} miles)")

                # Create backfill pull
                backfill_pull = self.create_pull_result(
                    warehouse=backfill_warehouse,
                    pull_qty=backfill_qty,
                    distance=int(backfill_dist),
                    total_available=backfill_avail,
                    equipment_description=equipment_description,
                    equipment_code=equipment_code,
                    df=df,
                    date_columns=date_columns,
                    is_backfill=True,
                    backfill_for_warehouse=warehouse
                )
                pulls.append(backfill_pull)

                return backfill_qty
            else:
                print(f"     ⚠️  No backfill source available")
                return 0

        return 0

    def _optimize_regional_strategy(self, df: pd.DataFrame, equipment_needs: Dict[str, int],
                                   destination: str, date_columns: List[str],
                                   warehouse_distances: Dict[str, float],
                                   selected_region: str) -> List[OptimizationResult]:
        """Regional scenario: Restrict pulls to warehouses within selected region"""
        results: List[OptimizationResult] = []
        
        distance_type = "road" if self.use_road_distances else "geodesic"
        region_warehouses = REGIONS[selected_region]['warehouses']
        
        print(f"Strategy: Regional optimization using {distance_type} distances (Region: {selected_region})")
        print(f"🌍 Restricting to region warehouses: {region_warehouses}")
        print("1. Check for partial availability (Qty Available > 0, Min Avail Qty < 0 = backfill needed)")
        print("2. Pull everything possible from arrival location first (even if outside region)")
        print("3. Arrange backfill from regional warehouses if arrival has partial availability")
        print("4. Pull from regional warehouses only")
        print("5. Use distance-based optimization within the region")
        
        # Create filtered warehouse distances for this region
        regional_distances = {}
        for warehouse, distance in warehouse_distances.items():
            if warehouse in region_warehouses:
                regional_distances[warehouse] = distance
        
        # Always allow pulling from destination even if not in region
        if destination not in regional_distances:
            regional_distances[destination] = 0.0
        
        print(f"🌍 Regional warehouses available: {list(regional_distances.keys())}")
        
        # Track inventory usage across all equipment to prevent double-pulling
        warehouse_inventory_used: Dict[str, Dict[str, int]] = {}
        
        # Initialize tracking for regional warehouses only
        for warehouse in regional_distances.keys():
            warehouse_inventory_used[warehouse] = {}
        
        # Process each equipment type independently
        for equipment_code, qty_needed in equipment_needs.items():
            equipment_description = self.get_equipment_description(df, equipment_code)
            weight_per_unit = self.weight_manager.get_weight(equipment_code)
            pulls: List[PullResult] = []
            remaining_need = qty_needed
            current_location = destination  # Start from arrival location
            
            print(f"\n--- REGIONAL ({selected_region}): Processing {equipment_code} (need {qty_needed}, weight: {weight_per_unit} lbs/unit) ---")
            
            # Initialize tracking for this equipment type
            for warehouse in warehouse_inventory_used.keys():
                if equipment_code not in warehouse_inventory_used[warehouse]:
                    warehouse_inventory_used[warehouse][equipment_code] = 0
            
            # Helper function to get REMAINING available quantity after previous pulls
            def get_remaining_quantity(warehouse: str, equipment: str) -> int:
                total_available = self.get_available_quantity(df, equipment, warehouse, date_columns)
                already_used = warehouse_inventory_used[warehouse].get(equipment, 0)
                remaining = max(0, total_available - already_used)
                print(f"    {warehouse} {equipment}: total={total_available}, used={already_used}, remaining={remaining}")
                return remaining
            
            # Step 1: Always pull maximum from arrival location first (even if outside region)
            if destination in warehouse_inventory_used:
                dest_remaining = get_remaining_quantity(destination, equipment_code)
                if dest_remaining > 0 and remaining_need > 0:
                    pull_qty = min(remaining_need, dest_remaining)
                    dest_total_available = self.get_available_quantity(df, equipment_code, destination, date_columns)

                    pull_result = self.create_pull_result(destination, pull_qty, 0, dest_total_available,
                                                        equipment_description, equipment_code, df, date_columns)
                    pulls.append(pull_result)
                    remaining_need -= pull_qty

                    # Track usage
                    warehouse_inventory_used[destination][equipment_code] += pull_qty

                    dest_region = self.get_warehouse_region(destination)
                    print(f"✅ {equipment_code}: pulled {pull_qty} from {destination} (ARRIVAL/LOCAL - {dest_region}) - remaining need: {remaining_need}, weight: {pull_result.total_weight} lbs")

                    # Check for partial availability and handle backfill
                    backfill_qty = self.handle_partial_availability(
                        df, equipment_code, equipment_description, destination,
                        destination, date_columns, regional_distances, pulls, selected_region
                    )
                    if backfill_qty > 0:
                        remaining_need -= backfill_qty
                        warehouse_inventory_used[pulls[-1].warehouse][equipment_code] += backfill_qty
            
            # Step 2: Regional optimization within selected region
            while remaining_need > 0:
                # Get all regional warehouses with REMAINING inventory
                warehouse_options = []
                current_coords = WAREHOUSE_COORDS[current_location]
                
                for warehouse in regional_distances.keys():
                    if warehouse == current_location:
                        continue  # Skip current location
                    
                    remaining_available = get_remaining_quantity(warehouse, equipment_code)
                    if remaining_available > 0:
                        # Calculate distance from CURRENT location to this warehouse
                        warehouse_coords = WAREHOUSE_COORDS[warehouse]
                        distance_from_current = self.calculate_distance(current_coords, warehouse_coords)
                        region = self.get_warehouse_region(warehouse)
                        warehouse_options.append((warehouse, remaining_available, distance_from_current, region))
                
                if not warehouse_options:
                    print(f"  ❌ No more regional warehouses have remaining {equipment_code}")
                    break
                
                print(f"  Regional warehouses with REMAINING inventory (from current location: {current_location}):")
                print(f"  Remaining need: {remaining_need}")
                for i, (wh, avail, dist, region) in enumerate(warehouse_options, 1):
                    fulfillment_pct = (avail / remaining_need) * 100 if remaining_need > 0 else 100
                    print(f"    {i}. {wh} ({region}): {avail} units ({fulfillment_pct:.0f}% of need) - {dist:.1f} miles from {current_location}")
                
                # Use same smart selection logic as standard optimization
                full_fulfillment_options = [(wh, avail, dist, region) for wh, avail, dist, region in warehouse_options if avail >= remaining_need]
                
                selected_warehouse = None
                
                if full_fulfillment_options:
                    # We have warehouses that can fulfill entire need - choose closest
                    full_fulfillment_options.sort(key=lambda x: x[2])  # Sort by distance
                    selected_warehouse = full_fulfillment_options[0]
                    warehouse, remaining_available, distance_from_current, region = selected_warehouse
                    print(f"  🎯 REGIONAL STRATEGY: Found regional warehouses that can fulfill entire need - choosing closest")
                    print(f"  🎯 SELECTED: {warehouse} ({region}) - can fulfill all {remaining_need} units, {distance_from_current:.1f} miles")
                    
                else:
                    # No warehouse can fulfill entire need
                    significant_options = [(wh, avail, dist, region) for wh, avail, dist, region in warehouse_options if (avail / remaining_need) >= 0.5]
                    
                    if significant_options:
                        # Use warehouses with significant quantities (≥50% of need)
                        significant_options.sort(key=lambda x: x[2])  # Sort by distance - closest first
                        selected_warehouse = significant_options[0]
                        warehouse, remaining_available, distance_from_current, region = selected_warehouse
                        print(f"  🎯 REGIONAL STRATEGY: No full fulfillment in region - choosing closest with significant quantity (≥50%)")
                        print(f"  🎯 SELECTED: {warehouse} ({region}) - {remaining_available} units = {(remaining_available/remaining_need)*100:.0f}% of need")
                        
                    else:
                        # All warehouses have small quantities - just take the closest
                        warehouse_options.sort(key=lambda x: x[2])  # Sort by distance
                        selected_warehouse = warehouse_options[0]
                        warehouse, remaining_available, distance_from_current, region = selected_warehouse
                        print(f"  🎯 REGIONAL STRATEGY: All regional warehouses have small quantities - taking closest")
                        print(f"  🎯 SELECTED: {warehouse} ({region}) - {remaining_available} units, closest regional option")
                
                if selected_warehouse is None:
                    break
                
                # Pull from selected warehouse
                pull_qty = min(remaining_need, remaining_available)
                
                # Calculate distance from original destination for reporting
                dest_coords = WAREHOUSE_COORDS[destination]
                warehouse_coords = WAREHOUSE_COORDS[warehouse]
                distance_from_dest = self.calculate_distance(dest_coords, warehouse_coords)
                
                # Get original total available for reporting
                total_available = self.get_available_quantity(df, equipment_code, warehouse, date_columns)

                pull_result = self.create_pull_result(warehouse, pull_qty, int(distance_from_dest),
                                                    total_available, equipment_description, equipment_code, df, date_columns)
                pulls.append(pull_result)
                remaining_need -= pull_qty

                # Track usage
                warehouse_inventory_used[warehouse][equipment_code] += pull_qty

                region = self.get_warehouse_region(warehouse)
                print(f"✅ {equipment_code}: pulled {pull_qty} from {warehouse} ({region}) ({distance_from_current:.1f} miles from {current_location}) - remaining need: {remaining_need}, weight: {pull_result.total_weight} lbs")

                # Check for partial availability and handle backfill
                backfill_qty = self.handle_partial_availability(
                    df, equipment_code, equipment_description, warehouse,
                    destination, date_columns, regional_distances, pulls, selected_region
                )
                if backfill_qty > 0:
                    remaining_need -= backfill_qty
                    warehouse_inventory_used[pulls[-1].warehouse][equipment_code] += backfill_qty

                # Update current location for next iteration
                current_location = warehouse
                print(f"  📍 Current location updated to: {current_location}")
                
                if remaining_need <= 0:
                    print(f"  ✓ {equipment_code} fully satisfied within {selected_region}!")
                    break
            
            # Calculate results for this equipment
            shortfall = max(0, remaining_need)
            if shortfall > 0:
                print(f"  ❌ {equipment_code}: shortfall of {shortfall} units (not available in {selected_region})")
            
            # Calculate total distance and weight
            equipment_distance = sum(pull.distance for pull in pulls if pull.distance > 0)
            equipment_total_weight = sum(pull.total_weight for pull in pulls)
            
            results.append(OptimizationResult(
                equipment_code=equipment_code,
                equipment_description=equipment_description,
                requested_quantity=qty_needed,
                pulls=pulls,
                shortfall=shortfall,
                total_distance=equipment_distance,
                weight_per_unit=weight_per_unit,
                total_weight=equipment_total_weight
            ))
        
        # Print final inventory usage summary for region
        print(f"\n=== FINAL REGIONAL INVENTORY USAGE SUMMARY ({selected_region}) ===")
        for warehouse in sorted(warehouse_inventory_used.keys()):
            used_items = {eq: qty for eq, qty in warehouse_inventory_used[warehouse].items() if qty > 0}
            if used_items:
                region = self.get_warehouse_region(warehouse)
                print(f"{warehouse} ({region}): {used_items}")
        
        return results

    def optimize_pulls_scenario(self, df: pd.DataFrame, equipment_needs: Dict[str, int],
                               destination: str, date_columns: List[str], 
                               preferred_source: Optional[str] = None,
                               scenario_name: str = "Standard",
                               region_filter: str = None) -> List[OptimizationResult]:
        """Run a single optimization scenario with optional regional filtering"""
        results: List[OptimizationResult] = []
        
        if destination not in WAREHOUSE_COORDS:
            raise ValueError(f"Destination '{destination}' not found in warehouse coordinates")
        
        all_distances = self.verify_warehouse_distances(destination)
        warehouse_distances = {wh: dist for wh, dist in all_distances.items() 
                             if wh != destination and dist < 900000}
        
        if not warehouse_distances:
            raise ValueError("No valid warehouses found for distance calculations")
        
        print(f"\n=== {scenario_name.upper()} OPTIMIZATION SCENARIO ===")
        if region_filter:
            print(f"🌍 REGIONAL FILTER: {region_filter}")
        
        # Different optimization strategies for different scenarios
        if scenario_name == "Standard":
            return self._optimize_standard_max_pull_strategy(
                df, equipment_needs, destination, date_columns, warehouse_distances
            )
        elif scenario_name == "Preferred Source":
            return self._optimize_preferred_source_strategy(
                df, equipment_needs, destination, date_columns, warehouse_distances, preferred_source
            )
        elif scenario_name == "Regional":
            return self._optimize_regional_strategy(
                df, equipment_needs, destination, date_columns, warehouse_distances, region_filter
            )
        else:  # Backup scenario
            return self._optimize_backup_strategy(
                df, equipment_needs, destination, date_columns, warehouse_distances
            )

    def _optimize_standard_max_pull_strategy(self, df: pd.DataFrame, equipment_needs: Dict[str, int],
                                           destination: str, date_columns: List[str], 
                                           warehouse_distances: Dict[str, float]) -> List[OptimizationResult]:
        """Standard scenario: Travel to closest warehouse with most inventory, then route from there"""
        results: List[OptimizationResult] = []
        
        distance_type = "road" if self.use_road_distances else "geodesic"
        print(f"Strategy: Routing optimization using {distance_type} distances - closest warehouse with most inventory, then route from there")
        print("1. Check for partial availability (Qty Available > 0, Min Avail Qty < 0 = backfill needed)")
        print("2. Pull everything possible from arrival location first")
        print("3. Arrange backfill if arrival location has partial availability")
        print("4. Find closest warehouse to arrival location that has MOST inventory")
        print("5. Pull all from that warehouse, then find closest warehouse to THAT location")
        print("6. Continue routing from last warehouse location until fulfilled")
        
        # Track inventory usage across all equipment to prevent double-pulling
        warehouse_inventory_used: Dict[str, Dict[str, int]] = {}
        
        # Initialize tracking for all warehouses
        for warehouse in WAREHOUSE_COORDS.keys():
            warehouse_inventory_used[warehouse] = {}
        
        # Process each equipment type independently
        for equipment_code, qty_needed in equipment_needs.items():
            equipment_description = self.get_equipment_description(df, equipment_code)
            weight_per_unit = self.weight_manager.get_weight(equipment_code)
            pulls: List[PullResult] = []
            remaining_need = qty_needed
            current_location = destination  # Start from arrival location
            
            print(f"\n--- STANDARD: Processing {equipment_code} (need {qty_needed}, weight: {weight_per_unit} lbs/unit) ---")
            
            # Initialize tracking for this equipment type
            for warehouse in warehouse_inventory_used.keys():
                if equipment_code not in warehouse_inventory_used[warehouse]:
                    warehouse_inventory_used[warehouse][equipment_code] = 0
            
            # Helper function to get REMAINING available quantity after previous pulls
            def get_remaining_quantity(warehouse: str, equipment: str) -> int:
                total_available = self.get_available_quantity(df, equipment, warehouse, date_columns)
                already_used = warehouse_inventory_used[warehouse].get(equipment, 0)
                remaining = max(0, total_available - already_used)
                print(f"    {warehouse} {equipment}: total={total_available}, used={already_used}, remaining={remaining}")
                return remaining
            
            # Step 1: Always pull maximum from arrival location first
            dest_remaining = get_remaining_quantity(destination, equipment_code)
            if dest_remaining > 0 and remaining_need > 0:
                pull_qty = min(remaining_need, dest_remaining)
                dest_total_available = self.get_available_quantity(df, equipment_code, destination, date_columns)

                pull_result = self.create_pull_result(destination, pull_qty, 0, dest_total_available,
                                                    equipment_description, equipment_code, df, date_columns)
                pulls.append(pull_result)
                remaining_need -= pull_qty

                # Track usage
                warehouse_inventory_used[destination][equipment_code] += pull_qty

                print(f"✅ {equipment_code}: pulled {pull_qty} from {destination} (ARRIVAL/LOCAL) - remaining need: {remaining_need}, weight: {pull_result.total_weight} lbs")

                # Check for partial availability and handle backfill
                backfill_qty = self.handle_partial_availability(
                    df, equipment_code, equipment_description, destination,
                    destination, date_columns, warehouse_distances, pulls, None
                )
                if backfill_qty > 0:
                    remaining_need -= backfill_qty
                    warehouse_inventory_used[pulls[-1].warehouse][equipment_code] += backfill_qty
            
            # Step 2: Smart routing - skip small quantities when better options exist
            while remaining_need > 0:
                # Get all warehouses with REMAINING inventory
                warehouse_options = []
                current_coords = WAREHOUSE_COORDS[current_location]
                
                for warehouse in WAREHOUSE_COORDS.keys():
                    if warehouse == current_location:
                        continue  # Skip current location
                    
                    remaining_available = get_remaining_quantity(warehouse, equipment_code)
                    if remaining_available > 0:
                        # Calculate distance from CURRENT location to this warehouse
                        warehouse_coords = WAREHOUSE_COORDS[warehouse]
                        distance_from_current = self.calculate_distance(current_coords, warehouse_coords)
                        warehouse_options.append((warehouse, remaining_available, distance_from_current))
                
                if not warehouse_options:
                    print(f"  ❌ No more warehouses have remaining {equipment_code}")
                    break
                
                print(f"  Warehouses with REMAINING inventory (from current location: {current_location}):")
                print(f"  Remaining need: {remaining_need}")
                for i, (wh, avail, dist) in enumerate(warehouse_options, 1):
                    fulfillment_pct = (avail / remaining_need) * 100 if remaining_need > 0 else 100
                    print(f"    {i}. {wh}: {avail} units ({fulfillment_pct:.0f}% of need) - {dist:.1f} miles from {current_location}")
                
                # SMART SELECTION LOGIC:
                # 1. If any warehouse can fulfill 100% of remaining need, choose the closest of those
                # 2. If no warehouse can fulfill 100%, avoid warehouses with <50% unless they're the only option
                # 3. Among viable options, choose the closest
                
                # Find warehouses that can fulfill entire remaining need (100%)
                full_fulfillment_options = [(wh, avail, dist) for wh, avail, dist in warehouse_options if avail >= remaining_need]
                
                selected_warehouse = None
                
                if full_fulfillment_options:
                    # We have warehouses that can fulfill entire need - choose closest
                    full_fulfillment_options.sort(key=lambda x: x[2])  # Sort by distance
                    selected_warehouse = full_fulfillment_options[0]
                    warehouse, remaining_available, distance_from_current = selected_warehouse
                    print(f"  🎯 STRATEGY: Found warehouses that can fulfill entire need - choosing closest")
                    print(f"  🎯 SELECTED: {warehouse} (can fulfill all {remaining_need} units, {distance_from_current:.1f} miles)")
                    
                else:
                    # No warehouse can fulfill entire need
                    # Filter out warehouses with very small quantities (< 50% of need) if better options exist
                    significant_options = [(wh, avail, dist) for wh, avail, dist in warehouse_options if (avail / remaining_need) >= 0.5]
                    
                    if significant_options:
                        # Use warehouses with significant quantities (≥50% of need)
                        significant_options.sort(key=lambda x: x[2])  # Sort by distance - closest first
                        selected_warehouse = significant_options[0]
                        warehouse, remaining_available, distance_from_current = selected_warehouse
                        print(f"  🎯 STRATEGY: No full fulfillment possible - choosing closest with significant quantity (≥50%)")
                        print(f"  🎯 SELECTED: {warehouse} ({remaining_available} units = {(remaining_available/remaining_need)*100:.0f}% of need)")
                        
                    else:
                        # All warehouses have small quantities - just take the closest
                        warehouse_options.sort(key=lambda x: x[2])  # Sort by distance
                        selected_warehouse = warehouse_options[0]
                        warehouse, remaining_available, distance_from_current = selected_warehouse
                        print(f"  🎯 STRATEGY: All warehouses have small quantities - taking closest")
                        print(f"  🎯 SELECTED: {warehouse} ({remaining_available} units, closest option)")
                
                if selected_warehouse is None:
                    break
                
                # Pull from selected warehouse
                pull_qty = min(remaining_need, remaining_available)
                
                # Calculate distance from original destination for reporting
                dest_coords = WAREHOUSE_COORDS[destination]
                warehouse_coords = WAREHOUSE_COORDS[warehouse]
                distance_from_dest = self.calculate_distance(dest_coords, warehouse_coords)
                
                # Get original total available for reporting
                total_available = self.get_available_quantity(df, equipment_code, warehouse, date_columns)

                pull_result = self.create_pull_result(warehouse, pull_qty, int(distance_from_dest),
                                                    total_available, equipment_description, equipment_code, df, date_columns)
                pulls.append(pull_result)
                remaining_need -= pull_qty

                # Track usage
                warehouse_inventory_used[warehouse][equipment_code] += pull_qty

                print(f"✅ {equipment_code}: pulled {pull_qty} from {warehouse} ({distance_from_current:.1f} miles from {current_location}) - remaining need: {remaining_need}, weight: {pull_result.total_weight} lbs")

                # Check for partial availability and handle backfill
                backfill_qty = self.handle_partial_availability(
                    df, equipment_code, equipment_description, warehouse,
                    destination, date_columns, warehouse_distances, pulls, None
                )
                if backfill_qty > 0:
                    remaining_need -= backfill_qty
                    warehouse_inventory_used[pulls[-1].warehouse][equipment_code] += backfill_qty

                # Update current location for next iteration
                current_location = warehouse
                print(f"  📍 Current location updated to: {current_location}")
                
                if remaining_need <= 0:
                    print(f"  ✓ {equipment_code} fully satisfied!")
                    break
            
            # Calculate results for this equipment
            shortfall = max(0, remaining_need)
            if shortfall > 0:
                print(f"  ❌ {equipment_code}: shortfall of {shortfall} units")
            
            # Calculate total distance and weight
            equipment_distance = sum(pull.distance for pull in pulls if pull.distance > 0)
            equipment_total_weight = sum(pull.total_weight for pull in pulls)
            
            results.append(OptimizationResult(
                equipment_code=equipment_code,
                equipment_description=equipment_description,
                requested_quantity=qty_needed,
                pulls=pulls,
                shortfall=shortfall,
                total_distance=equipment_distance,
                weight_per_unit=weight_per_unit,
                total_weight=equipment_total_weight
            ))
        
        # Print final inventory usage summary
        print(f"\n=== FINAL INVENTORY USAGE SUMMARY ===")
        for warehouse in sorted(warehouse_inventory_used.keys()):
            used_items = {eq: qty for eq, qty in warehouse_inventory_used[warehouse].items() if qty > 0}
            if used_items:
                print(f"{warehouse}: {used_items}")
        
        return results

    def _optimize_preferred_source_strategy(self, df: pd.DataFrame, equipment_needs: Dict[str, int],
                                          destination: str, date_columns: List[str], 
                                          warehouse_distances: Dict[str, float],
                                          preferred_source: str) -> List[OptimizationResult]:
        """Preferred source scenario: Prioritize preferred warehouse after destination"""
        results: List[OptimizationResult] = []
        
        distance_type = "road" if self.use_road_distances else "geodesic"
        print(f"Strategy: Preferred source optimization using {distance_type} distances (prioritizing {preferred_source})")
        print("1. Check for partial availability (Qty Available > 0, Min Avail Qty < 0 = backfill needed)")
        print("2. Pull everything possible from destination first")
        print("3. Arrange backfill if destination has partial availability")
        print(f"4. Pull maximum possible from preferred source: {preferred_source}")
        print("5. Arrange backfill if preferred source has partial availability")
        print("6. Use max-pull strategy for remaining needs")
        
        for equipment_code, qty_needed in equipment_needs.items():
            equipment_description = self.get_equipment_description(df, equipment_code)
            weight_per_unit = self.weight_manager.get_weight(equipment_code)
            pulls: List[PullResult] = []
            remaining_need = qty_needed
            
            print(f"\n--- PREFERRED SOURCE: Processing {equipment_code} (need {qty_needed}, weight: {weight_per_unit} lbs/unit) ---")
            
            # Step 1: Pull from destination first
            dest_available = self.get_available_quantity(df, equipment_code, destination, date_columns)
            if dest_available > 0 and remaining_need > 0:
                pull_qty = min(remaining_need, dest_available)
                pull_result = self.create_pull_result(destination, pull_qty, 0, dest_available,
                                                    equipment_description, equipment_code, df, date_columns)
                pulls.append(pull_result)
                remaining_need -= pull_qty
                print(f"✅ {equipment_code}: pulled {pull_qty} from {destination} (LOCAL) - remaining need: {remaining_need}, weight: {pull_result.total_weight} lbs")

                # Check for partial availability and handle backfill
                backfill_qty = self.handle_partial_availability(
                    df, equipment_code, equipment_description, destination,
                    destination, date_columns, warehouse_distances, pulls, None
                )
                if backfill_qty > 0:
                    remaining_need -= backfill_qty
            
            # Step 2: Pull from preferred source next
            if remaining_need > 0 and preferred_source in warehouse_distances:
                pref_available = self.get_available_quantity(df, equipment_code, preferred_source, date_columns)
                if pref_available > 0:
                    pull_qty = min(remaining_need, pref_available)
                    distance = int(warehouse_distances[preferred_source])
                    pull_result = self.create_pull_result(preferred_source, pull_qty, distance, pref_available,
                                                        equipment_description, equipment_code, df, date_columns)
                    pulls.append(pull_result)
                    remaining_need -= pull_qty
                    print(f"⭐ {equipment_code}: pulled {pull_qty} from {preferred_source} (PREFERRED) - remaining need: {remaining_need}, weight: {pull_result.total_weight} lbs")

                    # Check for partial availability and handle backfill
                    backfill_qty = self.handle_partial_availability(
                        df, equipment_code, equipment_description, preferred_source,
                        destination, date_columns, warehouse_distances, pulls, None
                    )
                    if backfill_qty > 0:
                        remaining_need -= backfill_qty
            
            # Step 3: Use max-pull strategy for remaining needs
            if remaining_need > 0:
                warehouse_options = []
                
                for warehouse, distance in warehouse_distances.items():
                    if warehouse == preferred_source:
                        continue  # Already handled
                    
                    available = self.get_available_quantity(df, equipment_code, warehouse, date_columns)
                    if available > 0:
                        warehouse_options.append((warehouse, available, distance))
                
                # Sort by quantity available (descending), then by distance (ascending)
                warehouse_options.sort(key=lambda x: (-x[1], x[2]))
                
                for warehouse, available, distance in warehouse_options:
                    if remaining_need <= 0:
                        break

                    pull_qty = min(remaining_need, available)
                    pull_result = self.create_pull_result(warehouse, pull_qty, int(distance), available,
                                                        equipment_description, equipment_code, df, date_columns)
                    pulls.append(pull_result)
                    remaining_need -= pull_qty
                    print(f"✅ {equipment_code}: pulled {pull_qty} from {warehouse} ({distance:.1f} miles) - remaining need: {remaining_need}, weight: {pull_result.total_weight} lbs")

                    # Check for partial availability and handle backfill
                    backfill_qty = self.handle_partial_availability(
                        df, equipment_code, equipment_description, warehouse,
                        destination, date_columns, warehouse_distances, pulls, None
                    )
                    if backfill_qty > 0:
                        remaining_need -= backfill_qty
                    
                    if remaining_need <= 0:
                        break
            
            shortfall = max(0, remaining_need)
            equipment_distance = sum(pull.distance for pull in pulls if pull.distance > 0)
            equipment_total_weight = sum(pull.total_weight for pull in pulls)
            
            results.append(OptimizationResult(
                equipment_code=equipment_code,
                equipment_description=equipment_description,
                requested_quantity=qty_needed,
                pulls=pulls,
                shortfall=shortfall,
                total_distance=equipment_distance,
                weight_per_unit=weight_per_unit,
                total_weight=equipment_total_weight
            ))
        
        return results

    def _optimize_backup_strategy(self, df: pd.DataFrame, equipment_needs: Dict[str, int],
                                destination: str, date_columns: List[str], 
                                warehouse_distances: Dict[str, float]) -> List[OptimizationResult]:
        """Backup scenario: Show all available options"""
        results: List[OptimizationResult] = []
        
        distance_type = "road" if self.use_road_distances else "geodesic"
        print(f"Strategy: Backup/alternative using {distance_type} distances showing all available options")
        print("1. Check for partial availability (Qty Available > 0, Min Avail Qty < 0 = backfill needed)")
        print("2. Show all warehouses with available inventory sorted by distance/availability")
        print("3. Arrange backfill for any warehouse with partial availability")
        
        for equipment_code, qty_needed in equipment_needs.items():
            equipment_description = self.get_equipment_description(df, equipment_code)
            weight_per_unit = self.weight_manager.get_weight(equipment_code)
            backup_pulls = []
            remaining_backup_need = qty_needed
            
            # Get all warehouse options sorted by best availability/distance score
            all_options = self.find_warehouse_options_for_equipment(
                df, equipment_code, destination, date_columns, qty_needed
            )
            
            # Create backup pulls from available options
            for warehouse, total_avail, distance, score in all_options:
                if total_avail > 0 and remaining_backup_need > 0:
                    pull_qty = min(remaining_backup_need, total_avail)
                    pull_result = self.create_pull_result(warehouse, pull_qty, int(distance), total_avail,
                                                        equipment_description, equipment_code, df, date_columns)
                    backup_pulls.append(pull_result)
                    remaining_backup_need -= pull_qty

                    # Check for partial availability and handle backfill
                    backfill_qty = self.handle_partial_availability(
                        df, equipment_code, equipment_description, warehouse,
                        destination, date_columns, warehouse_distances, backup_pulls, None
                    )
                    if backfill_qty > 0:
                        remaining_backup_need -= backfill_qty

                    if remaining_backup_need <= 0:
                        break
            
            shortfall = max(0, qty_needed - sum(p.quantity for p in backup_pulls))
            equipment_distance = 0
            for warehouse in set(pull.warehouse for pull in backup_pulls):
                if warehouse != destination:
                    dest_coords = WAREHOUSE_COORDS[destination]
                    wh_coords = WAREHOUSE_COORDS[warehouse]
                    equipment_distance += self.calculate_distance(wh_coords, dest_coords)
            
            equipment_total_weight = sum(pull.total_weight for pull in backup_pulls)
            
            results.append(OptimizationResult(
                equipment_code=equipment_code,
                equipment_description=equipment_description,
                requested_quantity=qty_needed,
                pulls=backup_pulls,
                shortfall=shortfall,
                total_distance=equipment_distance,
                weight_per_unit=weight_per_unit,
                total_weight=equipment_total_weight
            ))
        
        return results

    def optimize_pulls(self, df: pd.DataFrame, equipment_needs: Dict[str, int],
                      destination: str, date_columns: List[str], 
                      preferred_source: Optional[str] = None,
                      region_filter: Optional[str] = None) -> List[OptimizationScenario]:
        """Main optimization method - returns optimization scenarios with optional regional filtering"""
        scenarios = []
        
        distance_type = "road-based" if self.use_road_distances else "geodesic"
        print(f"\n{'='*80}")
        print(f"RUNNING OPTIMIZATION SCENARIOS ({distance_type.upper()} DISTANCES)")
        print(f"Destination: {destination}")
        print(f"Preferred Source: {preferred_source if preferred_source else 'None'}")
        print(f"Regional Filter: {region_filter if region_filter else 'None'}")
        print(f"Equipment Needs: {equipment_needs}")
        if self.use_road_distances:
            print("🛣️  Using OpenRouteService for realistic road distances")
        else:
            print("📏 Using geodesic distances (straight line)")
        print(f"{'='*80}")
        
        # Scenario 1: Regional Optimization (if region specified)
        if region_filter:
            print(f"\n🌍 SCENARIO 1: REGIONAL OPTIMIZATION ({region_filter})")
            regional_results = self.optimize_pulls_scenario(
                df, equipment_needs, destination, date_columns, 
                None, "Regional", region_filter
            )
            
            # Calculate metrics
            total_items = sum(r.requested_quantity for r in regional_results)
            total_shortfall = sum(r.shortfall for r in regional_results)
            fill_rate = ((total_items - total_shortfall) / total_items * 100) if total_items > 0 else 0
            total_weight = sum(r.total_weight for r in regional_results)
            
            warehouse_trips = {}
            for r in regional_results:
                for pull in r.pulls:
                    if pull.warehouse not in warehouse_trips:
                        warehouse_trips[pull.warehouse] = []
                    warehouse_trips[pull.warehouse].append(f"{r.equipment_code}({pull.quantity})")
            
            total_trips = len(warehouse_trips)
            consolidated_trips = sum(1 for trips in warehouse_trips.values() if len(trips) > 1)
            
            total_distance = 0
            for warehouse in warehouse_trips:
                if warehouse != destination:
                    dest_coords = WAREHOUSE_COORDS[destination]
                    wh_coords = WAREHOUSE_COORDS[warehouse]
                    trip_distance = self.calculate_distance(wh_coords, dest_coords)
                    total_distance += trip_distance
            
            scenarios.append(OptimizationScenario(
                name=f"Regional ({region_filter})",
                description=f"Regional optimization restricting to {region_filter}: {REGIONS[region_filter]['description']}",
                results=regional_results,
                total_distance=total_distance,
                fill_rate=fill_rate,
                total_trips=total_trips,
                consolidated_trips=consolidated_trips,
                total_weight=total_weight
            ))
        
        # Scenario 2: Preferred Source (if specified)
        if preferred_source:
            print(f"\n🎯 SCENARIO 2: PREFERRED SOURCE OPTIMIZATION")
            preferred_results = self.optimize_pulls_scenario(
                df, equipment_needs, destination, date_columns, 
                preferred_source, "Preferred Source"
            )
            
            # Calculate metrics
            total_items = sum(r.requested_quantity for r in preferred_results)
            total_shortfall = sum(r.shortfall for r in preferred_results)
            fill_rate = ((total_items - total_shortfall) / total_items * 100) if total_items > 0 else 0
            total_weight = sum(r.total_weight for r in preferred_results)
            
            warehouse_trips = {}
            for r in preferred_results:
                for pull in r.pulls:
                    if pull.warehouse not in warehouse_trips:
                        warehouse_trips[pull.warehouse] = []
                    warehouse_trips[pull.warehouse].append(f"{r.equipment_code}({pull.quantity})")
            
            total_trips = len(warehouse_trips)
            consolidated_trips = sum(1 for trips in warehouse_trips.values() if len(trips) > 1)
            
            total_distance = 0
            for warehouse in warehouse_trips:
                if warehouse != destination:
                    dest_coords = WAREHOUSE_COORDS[destination]
                    wh_coords = WAREHOUSE_COORDS[warehouse]
                    trip_distance = self.calculate_distance(wh_coords, dest_coords)
                    total_distance += trip_distance
            
            scenarios.append(OptimizationScenario(
                name="Preferred Source",
                description=f"Optimization prioritizing {preferred_source} after destination",
                results=preferred_results,
                total_distance=total_distance,
                fill_rate=fill_rate,
                total_trips=total_trips,
                consolidated_trips=consolidated_trips,
                total_weight=total_weight
            ))
        
        # Scenario 3: Standard Distance-Based Optimization
        scenario_number = len(scenarios) + 1
        print(f"\n📏 SCENARIO {scenario_number}: STANDARD DISTANCE-BASED OPTIMIZATION")
        standard_results = self.optimize_pulls_scenario(
            df, equipment_needs, destination, date_columns, 
            None, "Standard"
        )
        
        # Calculate metrics
        total_items = sum(r.requested_quantity for r in standard_results)
        total_shortfall = sum(r.shortfall for r in standard_results)
        fill_rate = ((total_items - total_shortfall) / total_items * 100) if total_items > 0 else 0
        total_weight = sum(r.total_weight for r in standard_results)
        
        warehouse_trips = {}
        for r in standard_results:
            for pull in r.pulls:
                if pull.warehouse not in warehouse_trips:
                    warehouse_trips[pull.warehouse] = []
                warehouse_trips[pull.warehouse].append(f"{r.equipment_code}({pull.quantity})")
        
        total_trips = len(warehouse_trips)
        consolidated_trips = sum(1 for trips in warehouse_trips.values() if len(trips) > 1)
        
        total_distance = 0
        for warehouse in warehouse_trips:
            if warehouse != destination:
                dest_coords = WAREHOUSE_COORDS[destination]
                wh_coords = WAREHOUSE_COORDS[warehouse]
                trip_distance = self.calculate_distance(wh_coords, dest_coords)
                total_distance += trip_distance
        
        scenarios.append(OptimizationScenario(
            name="Standard",
            description=f"Standard {distance_type} optimization with trip consolidation",
            results=standard_results,
            total_distance=total_distance,
            fill_rate=fill_rate,
            total_trips=total_trips,
            consolidated_trips=consolidated_trips,
            total_weight=total_weight
        ))
        
        # Scenario 4: Backup/Alternative Optimization
        scenario_number = len(scenarios) + 1
        print(f"\n🔄 SCENARIO {scenario_number}: BACKUP/ALTERNATIVE OPTIMIZATION")
        backup_results = []

        # Calculate warehouse distances for backfill detection
        all_distances = self.verify_warehouse_distances(destination)
        warehouse_distances = {wh: dist for wh, dist in all_distances.items()
                             if wh != destination and dist < 900000}
        
        for equipment_code, qty_needed in equipment_needs.items():
            equipment_description = self.get_equipment_description(df, equipment_code)
            weight_per_unit = self.weight_manager.get_weight(equipment_code)
            backup_pulls = []
            remaining_backup_need = qty_needed
            
            # Get all warehouse options sorted by best availability/distance score
            all_options = self.find_warehouse_options_for_equipment(
                df, equipment_code, destination, date_columns, qty_needed, region_filter
            )
            
            # Create backup pulls from available options
            for warehouse, total_avail, distance, score in all_options:
                if total_avail > 0 and remaining_backup_need > 0:
                    pull_qty = min(remaining_backup_need, total_avail)
                    pull_result = self.create_pull_result(warehouse, pull_qty, int(distance), total_avail,
                                                        equipment_description, equipment_code, df, date_columns)
                    backup_pulls.append(pull_result)
                    remaining_backup_need -= pull_qty

                    # Check for partial availability and handle backfill
                    backfill_qty = self.handle_partial_availability(
                        df, equipment_code, equipment_description, warehouse,
                        destination, date_columns, warehouse_distances, backup_pulls, None
                    )
                    if backfill_qty > 0:
                        remaining_backup_need -= backfill_qty

                    if remaining_backup_need <= 0:
                        break
            
            shortfall = max(0, qty_needed - sum(p.quantity for p in backup_pulls))
            equipment_trips = set(pull.warehouse for pull in backup_pulls if pull.quantity > 0)
            equipment_distance = 0
            for warehouse in equipment_trips:
                if warehouse != destination:
                    dest_coords = WAREHOUSE_COORDS[destination]
                    wh_coords = WAREHOUSE_COORDS[warehouse]
                    equipment_distance += self.calculate_distance(wh_coords, dest_coords)
            
            equipment_total_weight = sum(pull.total_weight for pull in backup_pulls)
            
            backup_results.append(OptimizationResult(
                equipment_code=equipment_code,
                equipment_description=equipment_description,
                requested_quantity=qty_needed,
                pulls=backup_pulls,
                shortfall=shortfall,
                total_distance=equipment_distance,
                weight_per_unit=weight_per_unit,
                total_weight=equipment_total_weight
            ))
        
        # Calculate backup metrics
        total_items = sum(r.requested_quantity for r in backup_results)
        total_shortfall = sum(r.shortfall for r in backup_results)
        fill_rate = ((total_items - total_shortfall) / total_items * 100) if total_items > 0 else 0
        total_weight = sum(r.total_weight for r in backup_results)
        
        warehouse_trips = {}
        for r in backup_results:
            for pull in r.pulls:
                if pull.warehouse not in warehouse_trips:
                    warehouse_trips[pull.warehouse] = []
                warehouse_trips[pull.warehouse].append(f"{r.equipment_code}({pull.quantity})")
        
        total_trips = len(warehouse_trips)
        consolidated_trips = sum(1 for trips in warehouse_trips.values() if len(trips) > 1)
        
        total_distance = 0
        for warehouse in warehouse_trips:
            if warehouse != destination:
                dest_coords = WAREHOUSE_COORDS[destination]
                wh_coords = WAREHOUSE_COORDS[warehouse]
                trip_distance = self.calculate_distance(wh_coords, dest_coords)
                total_distance += trip_distance
        
        backup_name = f"Backup ({region_filter})" if region_filter else "Backup"
        backup_desc = f"Alternative {distance_type} optimization showing all available options"
        if region_filter:
            backup_desc += f" within {region_filter}"
        
        scenarios.append(OptimizationScenario(
            name=backup_name,
            description=backup_desc,
            results=backup_results,
            total_distance=total_distance,
            fill_rate=fill_rate,
            total_trips=total_trips,
            consolidated_trips=consolidated_trips,
            total_weight=total_weight
        ))
        
        print(f"\n{'='*80}")
        print(f"COMPLETED ALL SCENARIOS ({distance_type.upper()} DISTANCES)")
        print(f"Scenarios generated: {[s.name for s in scenarios]}")
        print(f"{'='*80}")
        
        return scenarios


class ConfigManager:
    """Handle configuration persistence"""
    
    @staticmethod
    def save_config(data: dict):
        try:
            with open(CONFIG_FILE, 'w') as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            print(f"Warning: Could not save config: {e}")
    
    @staticmethod
    def load_config() -> dict:
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r') as f:
                    return json.load(f)
        except Exception as e:
            print(f"Warning: Could not load config: {e}")
        return {}


class LogisticsManagementSuite:
    def __init__(self, root):
        self.root = root
        self.root.title("Logistics Management Suite v4.3 - With Regional Optimization & OpenRouteService")
        self.root.geometry("1500x1000")
        
        # Initialize components
        self.engine = LogisticsEngine()
        self.weight_manager = self.engine.weight_manager
        self.config = ConfigManager.load_config()
        
        # Create main notebook for tabs
        self.main_notebook = ttk.Notebook(self.root)
        self.main_notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Create tabs
        self.create_quote_converter_tab()
        self.create_optimizer_tab()
        self.create_weight_management_tab()
        self.create_settings_tab()
        
        # Initialize optimizer state
        self.init_optimizer_state()
        self.load_saved_settings()
    
    def create_quote_converter_tab(self):
        """Create the Quote Converter tab"""
        self.quote_frame = ttk.Frame(self.main_notebook)
        self.main_notebook.add(self.quote_frame, text="📄 Quote Converter")
        
        # Variables for quote converter
        self.input_file = tk.StringVar()
        self.output_file = tk.StringVar()
        
        # Main frame with padding
        main_frame = ttk.Frame(self.quote_frame, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="Rental Quote Format Converter", 
                               font=('Arial', 16, 'bold'))
        title_label.pack(pady=(0, 20))
        
        # Description
        desc_text = ("This tool converts detailed rental quotes (47+ columns) to simplified format (4 columns).\n"
                    "Output format: Main | Equipment | Description | Ordered")
        desc_label = ttk.Label(main_frame, text=desc_text, font=('Arial', 10))
        desc_label.pack(pady=(0, 20))
        
        # File selection frame
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="10")
        file_frame.pack(fill='x', pady=10)
        
        # Input file selection
        input_frame = tk.Frame(file_frame)
        input_frame.pack(fill='x', pady=5)
        ttk.Label(input_frame, text="Input Excel File:", width=15).pack(side=tk.LEFT)
        ttk.Entry(input_frame, textvariable=self.input_file, width=60).pack(side=tk.LEFT, padx=(5, 5), fill=tk.X, expand=True)
        ttk.Button(input_frame, text="Browse", command=self.browse_input_file).pack(side=tk.RIGHT)
        
        # Output file selection
        output_frame = tk.Frame(file_frame)
        output_frame.pack(fill='x', pady=5)
        ttk.Label(output_frame, text="Output Excel File:", width=15).pack(side=tk.LEFT)
        ttk.Entry(output_frame, textvariable=self.output_file, width=60).pack(side=tk.LEFT, padx=(5, 5), fill=tk.X, expand=True)
        ttk.Button(output_frame, text="Browse", command=self.browse_output_file).pack(side=tk.RIGHT)
        
        # Convert button and progress
        control_frame = tk.Frame(main_frame)
        control_frame.pack(fill='x', pady=20)
        
        self.convert_btn = ttk.Button(control_frame, text="🔄 Convert Quote Format", 
                                     command=self.convert_quote)
        self.convert_btn.pack()
        
        # Progress bar
        self.quote_progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.quote_progress.pack(fill='x', pady=5)
        
        # Status label
        self.quote_status_label = ttk.Label(main_frame, text="Ready to convert", foreground="blue")
        self.quote_status_label.pack(pady=5)
        
        # Results text area
        results_frame = ttk.LabelFrame(main_frame, text="Conversion Results", padding="5")
        results_frame.pack(fill='both', expand=True, pady=10)
        
        # Create text widget with scrollbar
        text_frame = tk.Frame(results_frame)
        text_frame.pack(fill='both', expand=True)
        
        self.quote_results_text = tk.Text(text_frame, height=15, width=80, wrap=tk.WORD)
        quote_scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=self.quote_results_text.yview)
        self.quote_results_text.configure(yscrollcommand=quote_scrollbar.set)
        
        self.quote_results_text.pack(side="left", fill="both", expand=True)
        quote_scrollbar.pack(side="right", fill="y")
    
    def create_optimizer_tab(self):
        """Create the Warehouse Optimizer tab with regional functionality"""
        self.optimizer_frame = ttk.Frame(self.main_notebook)
        self.main_notebook.add(self.optimizer_frame, text="🌍 Warehouse Optimizer")
        
        # Create notebook for optimizer sub-tabs
        self.optimizer_notebook = ttk.Notebook(self.optimizer_frame)
        self.optimizer_notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Main optimization tab
        self.opt_main_frame = ttk.Frame(self.optimizer_notebook)
        self.optimizer_notebook.add(self.opt_main_frame, text="Optimization Dashboard")
        
        self.create_optimizer_main_tab()
        self.create_diagnostic_tab()
    
    def create_diagnostic_tab(self):
        """Create the Diagnostic Tab for inspecting single items"""
        self.diag_frame = ttk.Frame(self.optimizer_notebook)
        self.optimizer_notebook.add(self.diag_frame, text="🔍 Item Diagnostics")
        
        # Main container
        main_frame = ttk.Frame(self.diag_frame, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # --- Input Section ---
        input_frame = ttk.LabelFrame(main_frame, text="Diagnostic Query", padding="10")
        input_frame.pack(fill='x', pady=5)
        
        # Item Code
        tk.Label(input_frame, text="Item Code:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky='e', padx=5, pady=5)
        self.diag_item_var = tk.StringVar()
        entry = ttk.Entry(input_frame, textvariable=self.diag_item_var, width=25)
        entry.grid(row=0, column=1, sticky='w', padx=5, pady=5)
        
        # Location
        tk.Label(input_frame, text="Target Warehouse:", font=("Arial", 10, "bold")).grid(row=0, column=2, sticky='e', padx=5, pady=5)
        self.diag_loc_var = tk.StringVar()
        loc_cb = ttk.Combobox(input_frame, textvariable=self.diag_loc_var, values=list(WAREHOUSE_COORDS.keys()), state="readonly", width=20)
        loc_cb.grid(row=0, column=3, sticky='w', padx=5, pady=5)
        if WAREHOUSE_COORDS:
            loc_cb.current(0)
            
        # Analyze Button
        btn = ttk.Button(input_frame, text="🔍 Analyze Item Availability", command=self.run_diagnostic_analysis)
        btn.grid(row=0, column=4, padx=20, pady=5)
        
        # --- Results Section ---
        results_frame = ttk.LabelFrame(main_frame, text="Diagnostic Results", padding="10")
        results_frame.pack(fill='both', expand=True, pady=10)
        
        # We'll use a text widget for rich output
        self.diag_text = tk.Text(results_frame, height=20, width=80, font=("Consolas", 10))
        scroll = ttk.Scrollbar(results_frame, orient="vertical", command=self.diag_text.yview)
        self.diag_text.configure(yscrollcommand=scroll.set)
        
        self.diag_text.pack(side='left', fill='both', expand=True)
        scroll.pack(side='right', fill='y')
        
        # Instructions
        self.diag_text.insert('1.0', "Enter an Equipment Code and Select a Warehouse above, then click Analyze.\n")
        self.diag_text.insert('end', "This will show you exactly what the script sees in the Excel file.")

    def run_diagnostic_analysis(self):
        """Run the diagnostic query"""
        item_code = self.diag_item_var.get().strip()
        location = self.diag_loc_var.get()
        
        if not item_code or not location:
            messagebox.showwarning("Input Missing", "Please enter both Item Code and Location.")
            return
            
        if self.df is None:
            messagebox.showerror("No Data", "Please load an Inventory File first in the Optimization tab.")
            return

        self.diag_text.delete('1.0', 'end')
        self.diag_text.insert('end', f"=== DIAGNOSTIC REPORT: {item_code} @ {location} ===\n\n")
        
        # 1. Raw Data Lookup
        self.diag_text.insert('end', "1. RAW DATA LOOKUP\n")
        self.diag_text.insert('end', "-" * 40 + "\n")
        
        # Check Qty Available
        qty_avail = self.engine.get_available_quantity(self.df, item_code, location, self.selected_date_columns)
        
        # Check Min Avail Qty
        min_avail = self.engine.get_min_avail_qty(self.df, item_code, location)
        
        self.diag_text.insert('end', f"  • Raw Qty Available:  {qty_avail}\n")
        self.diag_text.insert('end', f"  • Raw Min Avail Qty:  {min_avail}\n\n")
        
        # 2. Logic Interpretation
        self.diag_text.insert('end', "2. LOGIC INTERPRETATION\n")
        self.diag_text.insert('end', "-" * 40 + "\n")
        
        # Logic Scenario Check
        backfill_needed = 0
        scenario = "Standard Availability"
        
        if qty_avail > 0:
            if min_avail < 0:
                scenario = "NEGATIVE MIN AVAIL (Backfill Trigger)"
                backfill_needed = abs(min_avail)
                notes = "Local pull allowed, but backfill triggered by negative Min Avail."
            elif min_avail > qty_avail and min_avail > 0:
                scenario = "POSITIVE MIN AVAIL (Shortfall)"
                backfill_needed = min_avail - qty_avail
                notes = "Legacy logic: Total required is greater than available."
            else:
                notes = "Normal availability. No backfill needed."
        else:
            scenario = "NO AVAILABILITY"
            notes = "Location has 0 available. Skipped."
            
        self.diag_text.insert('end', f"  • Scenario Detected:  {scenario}\n")
        self.diag_text.insert('end', f"  • Action:             {notes}\n")
        if backfill_needed > 0:
            self.diag_text.insert('end', f"  • Backfill Needed:    {backfill_needed} units\n")
            
        self.diag_text.insert('end', "\n")
        
        # 3. Backfill Analysis
        if backfill_needed > 0 or qty_avail == 0:
            self.diag_text.insert('end', "3. POTENTIAL SOURCING OPTIONS (Backfill/Alternatives)\n")
            self.diag_text.insert('end', "-" * 40 + "\n")
            
            # Find options
            options = self.engine.find_warehouse_options_for_equipment(
                self.df, item_code, location, self.selected_date_columns, 1
            )
            
            # Filter out the target location itself
            options = [o for o in options if o[0] != location]
            
            if not options:
                self.diag_text.insert('end', "  ❌ No other warehouses found with inventory.\n")
            else:
                self.diag_text.insert('end', f"  {'Warehouse':<15} | {'Qty':<5} | {'Dist (mi)':<10} | {'Score':<10}\n")
                self.diag_text.insert('end', f"  {'-'*15}-+-{'-'*5}-+-{'-'*10}-+-{'-'*10}\n")
                
                for i, (wh, qty, dist, score) in enumerate(options[:5], 1):
                    mark = ""
                    if i == 1: mark = " (Likely Pick)"
                    self.diag_text.insert('end', f"  {wh:<15} | {qty:<5} | {dist:<10.1f} | {score:<10.2f}{mark}\n")

    def create_optimizer_main_tab(self):
        """Create the main optimization interface with REDESIGNED LAYOUT"""
        
        # Use PanedWindow for split view (Top controls / Bottom results)
        self.main_paned = ttk.PanedWindow(self.opt_main_frame, orient=tk.VERTICAL)
        self.main_paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # --- TOP SECTION: Configuration ---
        top_frame = ttk.Frame(self.main_paned)
        self.main_paned.add(top_frame, weight=1)
        
        # Row 1: File Loading & Status (Compact)
        file_frame = ttk.Frame(top_frame)
        file_frame.pack(fill='x', pady=5)
        
        # Buttons
        ttk.Button(file_frame, text="📂 Load Inventory File", command=self.load_inventory_file).pack(side='left', padx=5)
        ttk.Button(file_frame, text="Load Excel Quote (Optional)", command=self.load_needs_excel).pack(side='left', padx=5)
        
        # File Label
        self.file_label = ttk.Label(file_frame, text="No file loaded", font=("Arial", 9, "italic"))
        self.file_label.pack(side='left', padx=5)
        
        # Distance status indicator (Compact)
        dist_color = "green" if self.engine.use_road_distances else "orange"
        dist_text = "🛣️ Road Distances" if self.engine.use_road_distances else "📏 Geodesic"
        ttk.Label(file_frame, text=f"[{dist_text}]", foreground=dist_color).pack(side='right', padx=10)
        
        # Progress Bar (Restored)
        self.opt_progress = ttk.Progressbar(file_frame, mode='indeterminate', length=100)
        self.opt_progress.pack(side='right', padx=5)
        
        # Row 2: Setup (Left) and Equipment (Right)
        config_grid = ttk.Frame(top_frame)
        config_grid.pack(fill='both', expand=True, pady=5)
        
        # Left Column: Setup
        setup_frame = ttk.LabelFrame(config_grid, text="Logistics Configuration", padding="10")
        setup_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))
        
        # Date Selection
        date_row = ttk.Frame(setup_frame)
        date_row.pack(fill='x', pady=2)
        ttk.Label(date_row, text="Dates:").pack(side='left')
        ttk.Button(date_row, text="Select Columns", command=self.select_dates).pack(side='left', padx=5)
        self.dates_label = ttk.Label(date_row, text="None", width=15)
        self.dates_label.pack(side='left')
        
        # Separator
        ttk.Separator(setup_frame, orient='horizontal').pack(fill='x', pady=8)
        
        # Locations
        grid_loc = ttk.Frame(setup_frame)
        grid_loc.pack(fill='x')
        
        ttk.Label(grid_loc, text="Arrival Location:").grid(row=0, column=0, sticky='w', pady=2)
        self.destination_var = tk.StringVar()
        dest_cb = ttk.Combobox(grid_loc, textvariable=self.destination_var, values=list(WAREHOUSE_COORDS.keys()), state="readonly", width=18)
        dest_cb.grid(row=0, column=1, sticky='e', pady=2)
        
        ttk.Label(grid_loc, text="Regional Filter:").grid(row=1, column=0, sticky='w', pady=2)
        self.region_var = tk.StringVar()
        reg_cb = ttk.Combobox(grid_loc, textvariable=self.region_var, values=["(None - Use All Warehouses)"] + list(REGIONS.keys()), state="readonly", width=18)
        reg_cb.grid(row=1, column=1, sticky='e', pady=2)
        reg_cb.current(0)
        
        ttk.Label(grid_loc, text="Preferred Source:").grid(row=2, column=0, sticky='w', pady=2)
        self.preferred_source_var = tk.StringVar()
        pref_cb = ttk.Combobox(grid_loc, textvariable=self.preferred_source_var, values=["(None - Use Standard)"] + list(WAREHOUSE_COORDS.keys()), state="readonly", width=18)
        pref_cb.grid(row=2, column=1, sticky='e', pady=2)
        pref_cb.current(0)
        
        # Right Column: Equipment
        eq_frame = ttk.LabelFrame(config_grid, text="Equipment Needs", padding="5")
        eq_frame.pack(side='right', fill='both', expand=True, padx=(5, 0))
        
        # Eq Controls (Compact)
        eq_ctrl = ttk.Frame(eq_frame)
        eq_ctrl.pack(fill='x')
        ttk.Button(eq_ctrl, text="Set All 1", command=lambda: self.set_all_quantities(1), width=8).pack(side='left')
        ttk.Button(eq_ctrl, text="Clear", command=self.clear_all_quantities, width=8).pack(side='left', padx=2)
        
        # Scrollable List
        canv_frame = ttk.Frame(eq_frame)
        canv_frame.pack(fill='both', expand=True, pady=5)
        
        eq_canvas = tk.Canvas(canv_frame, height=120) # Fixed height to save space
        eq_scroll = ttk.Scrollbar(canv_frame, orient="vertical", command=eq_canvas.yview)
        self.eq_list_frame = ttk.Frame(eq_canvas)
        
        self.eq_list_frame.bind("<Configure>", lambda e: eq_canvas.configure(scrollregion=eq_canvas.bbox("all")))
        eq_canvas.create_window((0, 0), window=self.eq_list_frame, anchor="nw")
        eq_canvas.configure(yscrollcommand=eq_scroll.set)
        
        eq_canvas.pack(side="left", fill="both", expand=True)
        eq_scroll.pack(side="right", fill="y")
        
        # Row 3: Optimize Button
        self.optimize_button = ttk.Button(top_frame, text="🚀 RUN OPTIMIZATION 🚀", command=self.run_optimizer_threaded)
        self.optimize_button.pack(fill='x', pady=10, ipady=5)
        
        # --- BOTTOM SECTION: Results ---
        results_frame = ttk.LabelFrame(self.main_paned, text="Optimization Results")
        self.main_paned.add(results_frame, weight=3) # Give more weight to results
        
        # Results tools
        res_ctrl = ttk.Frame(results_frame)
        res_ctrl.pack(fill='x', pady=2)
        self.export_button = ttk.Button(res_ctrl, text="📊 Export to Excel", command=self.export_to_excel, state="disabled")
        self.export_button.pack(side='right', padx=5)
        
        # Notebook for text output
        self.results_notebook = ttk.Notebook(results_frame)
        self.results_notebook.pack(fill='both', expand=True)
        
        # Summary Tab
        summary_frame = ttk.Frame(self.results_notebook)
        self.results_notebook.add(summary_frame, text="Summary")
        self.summary_text = tk.Text(summary_frame, wrap=tk.WORD, font=("Arial", 10))
        sum_scroll = ttk.Scrollbar(summary_frame, command=self.summary_text.yview)
        self.summary_text.configure(yscrollcommand=sum_scroll.set)
        self.summary_text.pack(side='left', fill='both', expand=True)
        sum_scroll.pack(side='right', fill='y')
        
        # Detail Tab
        detail_frame = ttk.Frame(self.results_notebook)
        self.results_notebook.add(detail_frame, text="Detailed Breakdown")
        self.detail_text = tk.Text(detail_frame, wrap=tk.NONE, font=("Consolas", 9))
        det_scroll_y = ttk.Scrollbar(detail_frame, command=self.detail_text.yview)
        det_scroll_x = ttk.Scrollbar(detail_frame, orient='horizontal', command=self.detail_text.xview)
        self.detail_text.configure(yscrollcommand=det_scroll_y.set, xscrollcommand=det_scroll_x.set)
        
        self.detail_text.pack(side='top', fill='both', expand=True)
        det_scroll_x.pack(side='bottom', fill='x')
        det_scroll_y.pack(side='right', fill='y', before=self.detail_text) # Pack Y scrollbar appropriately
    
    def create_weight_management_tab(self):
        """Create the Equipment Weight Management tab"""
        self.weight_frame = ttk.Frame(self.main_notebook)
        self.main_notebook.add(self.weight_frame, text="⚖️ Equipment Weights")
        
        # Main frame with padding
        main_frame = ttk.Frame(self.weight_frame, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="Equipment Weight Management", 
                               font=('Arial', 16, 'bold'))
        title_label.pack(pady=(0, 10))
        
        # Description
        desc_text = ("Manage equipment weights (in pounds) for accurate load calculations.\n"
                    f"Weight data is stored in: {WEIGHTS_FILE}")
        desc_label = ttk.Label(main_frame, text=desc_text, font=('Arial', 10))
        desc_label.pack(pady=(0, 20))
        
        # Control buttons
        control_frame = tk.Frame(main_frame)
        control_frame.pack(fill='x', pady=10)
        
        ttk.Button(control_frame, text="📁 Load from CSV", 
                  command=self.load_weights_from_csv).pack(side="left", padx=5)
        ttk.Button(control_frame, text="💾 Export to CSV", 
                  command=self.export_weights_to_csv).pack(side="left", padx=5)
        ttk.Button(control_frame, text="🔄 Refresh List", 
                  command=self.refresh_weight_list).pack(side="left", padx=5)
        
        # Search frame
        search_frame = tk.Frame(main_frame)
        search_frame.pack(fill='x', pady=5)
        
        tk.Label(search_frame, text="Search Equipment:").pack(side="left", padx=5)
        self.weight_search_var = tk.StringVar()
        self.weight_search_var.trace_add('write', self.filter_weight_list)
        search_entry = ttk.Entry(search_frame, textvariable=self.weight_search_var, width=30)
        search_entry.pack(side="left", padx=5)
        
        # Weight list frame
        list_frame = ttk.LabelFrame(main_frame, text="Equipment Weights (lbs)", padding="5")
        list_frame.pack(fill='both', expand=True, pady=10)
        
        # Create treeview for equipment weights
        columns = ('Equipment', 'Weight (lbs)', 'Status')
        self.weight_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        
        # Configure columns
        self.weight_tree.heading('Equipment', text='Equipment Code')
        self.weight_tree.heading('Weight (lbs)', text='Weight (lbs)')
        self.weight_tree.heading('Status', text='Status')
        
        self.weight_tree.column('Equipment', width=200)
        self.weight_tree.column('Weight (lbs)', width=120)
        self.weight_tree.column('Status', width=100)
        
        # Scrollbar for treeview
        weight_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.weight_tree.yview)
        self.weight_tree.configure(yscrollcommand=weight_scroll.set)
        
        self.weight_tree.pack(side="left", fill="both", expand=True)
        weight_scroll.pack(side="right", fill="y")
        
        # Bind double-click to edit
        self.weight_tree.bind("<Double-1>", self.edit_weight)
        
        # Add/Edit frame
        edit_frame = ttk.LabelFrame(main_frame, text="Add/Edit Equipment Weight", padding="10")
        edit_frame.pack(fill='x', pady=10)
        
        edit_controls = tk.Frame(edit_frame)
        edit_controls.pack()
        
        tk.Label(edit_controls, text="Equipment Code:").grid(row=0, column=0, padx=5, pady=5, sticky='e')
        self.edit_equipment_var = tk.StringVar()
        ttk.Entry(edit_controls, textvariable=self.edit_equipment_var, width=20).grid(row=0, column=1, padx=5, pady=5)
        
        tk.Label(edit_controls, text="Weight (lbs):").grid(row=0, column=2, padx=5, pady=5, sticky='e')
        self.edit_weight_var = tk.StringVar()
        ttk.Entry(edit_controls, textvariable=self.edit_weight_var, width=15).grid(row=0, column=3, padx=5, pady=5)
        
        ttk.Button(edit_controls, text="💾 Save Weight", 
                  command=self.save_equipment_weight).grid(row=0, column=4, padx=10, pady=5)
        ttk.Button(edit_controls, text="🗑️ Delete", 
                  command=self.delete_equipment_weight).grid(row=0, column=5, padx=5, pady=5)
        
        # Status label
        self.weight_status_label = ttk.Label(edit_frame, text="Ready", foreground="blue")
        self.weight_status_label.pack(pady=5)
        
        # Load initial weight list
        self.refresh_weight_list()
    
    def create_settings_tab(self):
        """Create the Settings tab with regional information"""
        self.settings_frame = ttk.Frame(self.main_notebook)
        self.main_notebook.add(self.settings_frame, text="⚙️ Settings")
        
        # Regional Information Section
        regional_frame = ttk.LabelFrame(self.settings_frame, text="Regional Configuration")
        regional_frame.pack(fill='x', padx=5, pady=5)
        
        regional_info = """REGIONAL WAREHOUSE DEFINITIONS:

🌍 Region 1 (East/South):
   • CHICAGO (41.88°N, -87.63°W)
   • BOSTON (42.36°N, -71.06°W)  
   • PHILADELPHIA (39.95°N, -75.17°W)
   • NEW YORK (40.71°N, -74.01°W)
   • ORLANDO (28.54°N, -81.38°W)
   • FT LAUDERDALE (26.12°N, -80.14°W)
   • WDC (38.91°N, -77.04°W)
   • NASHVILLE (36.16°N, -86.78°W)

🌍 Region 2 (West/Central):
   • ANAHEIM (33.84°N, -117.91°W)
   • SAN FRANCISCO (37.77°N, -122.42°W)
   • PHOENIX (33.45°N, -112.07°W)
   • LAS VEGAS (36.17°N, -115.14°W)
   • DALLAS (32.78°N, -96.80°W)

Regional optimization restricts equipment pulls to warehouses within the selected region,
while still allowing pulls from the destination warehouse even if it's outside the region."""
        
        ttk.Label(regional_frame, text=regional_info, font=('Arial', 9), justify='left').pack(padx=10, pady=10)
        
        # OpenRouteService Status
        ors_frame = ttk.LabelFrame(self.settings_frame, text="OpenRouteService Configuration")
        ors_frame.pack(fill='x', padx=5, pady=5)
        
        if self.engine.use_road_distances:
            status_text = "✅ OpenRouteService CONNECTED - Using road-based distances"
            status_color = "green"
        else:
            status_text = "⚠️ OpenRouteService NOT CONNECTED - Using geodesic distances"
            status_color = "orange"
        
        ttk.Label(ors_frame, text=status_text, foreground=status_color, font=('Arial', 10, 'bold')).pack(pady=5)
        
        if not self.engine.use_road_distances:
            instruction_text = ("To enable road distances:\n"
                              "1. Sign up for free API key at: https://openrouteservice.org/dev/#/signup\n"
                              "2. Replace ORS_API_KEY in the code with your actual API key\n"
                              "3. Restart the application")
            ttk.Label(ors_frame, text=instruction_text, font=('Arial', 9), wraplength=500).pack(padx=10, pady=5)
        else:
            info_text = ("Free tier: 2000 requests per day\n"
                        "Road distances provide 20-50% more accurate logistics planning")
            ttk.Label(ors_frame, text=info_text, font=('Arial', 9), foreground="green").pack(padx=10, pady=5)
        
        # Distance verification section
        dist_frame = ttk.LabelFrame(self.settings_frame, text="Distance Verification")
        dist_frame.pack(fill='x', padx=5, pady=5)
        
        dist_control = tk.Frame(dist_frame)
        dist_control.pack(pady=5)
        
        tk.Label(dist_control, text="Verify distances from:").pack(side="left", padx=5)
        self.dist_destination_var = tk.StringVar()
        dist_dropdown = ttk.Combobox(dist_control, textvariable=self.dist_destination_var,
                                    values=list(WAREHOUSE_COORDS.keys()), state="readonly", width=15)
        dist_dropdown.pack(side="left", padx=5)
        dist_dropdown.current(0)
        
        verify_btn = ttk.Button(dist_control, text="🛣️ Verify Road Distances" if self.engine.use_road_distances else "📏 Verify Distances", 
                              command=self.verify_distances)
        verify_btn.pack(side="left", padx=5)
        
        # Distance results display
        self.distance_text = tk.Text(dist_frame, height=8, wrap=tk.NONE, font=("Consolas", 9))
        dist_scroll = ttk.Scrollbar(dist_frame, orient="vertical", command=self.distance_text.yview)
        self.distance_text.configure(yscrollcommand=dist_scroll.set)
        self.distance_text.pack(side="left", fill="both", expand=True)
        dist_scroll.pack(side="right", fill="y")
        
        # Warehouse management
        wh_frame = ttk.LabelFrame(self.settings_frame, text="Warehouse Coordinates")
        wh_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Warehouse list
        columns = ('Name', 'Latitude', 'Longitude')
        self.warehouse_tree = ttk.Treeview(wh_frame, columns=columns, show='headings', height=10)
        
        for col in columns:
            self.warehouse_tree.heading(col, text=col)
            self.warehouse_tree.column(col, width=120)
        
        wh_scroll = ttk.Scrollbar(wh_frame, orient="vertical", command=self.warehouse_tree.yview)
        self.warehouse_tree.configure(yscrollcommand=wh_scroll.set)
        
        self.warehouse_tree.pack(side="left", fill="both", expand=True)
        wh_scroll.pack(side="right", fill="y")
        
        self.populate_warehouse_tree()
        
        # About section with regional features
        about_frame = ttk.LabelFrame(self.settings_frame, text="About")
        about_frame.pack(fill='x', padx=5, pady=5)
        
        about_text = """Logistics Management Suite v4.3 - With Regional Optimization & OpenRouteService Road Distances
        
NEW in v4.3:
• 🌍 Regional Optimization: Restrict equipment pulls to specific geographical regions
• 🌍 Region 1 (East/South): 8 warehouses covering Eastern and Southern US
• 🌍 Region 2 (West/Central): 5 warehouses covering Western and Central US
• 🌍 Smart Regional Logic: Always allows pulls from destination even if outside selected region
• 🌍 Regional Scenario Generation: Creates dedicated regional optimization scenarios
• 🌍 Enhanced Reporting: Regional information included in all exports and displays

Previous v4.2 Features:
• 🛣️ Road-Based Distance Calculation: Uses OpenRouteService API for realistic driving distances
• 📏 20-50% More Accurate: Road distances vs straight-line for better logistics planning
• 🌍 Global Coverage: Worldwide road network data via OpenRouteService
• 🔄 Automatic Fallback: Falls back to geodesic if API unavailable
• ⚡ Cached Results: Avoids repeated API calls for same route calculations
• 🎯 Enhanced Optimization: More realistic warehouse selection based on actual travel

Regional Benefits:
• Compliance with regional service agreements
• Reduced cross-country shipping costs
• Better load balancing within geographical areas
• Simplified logistics coordination within regions
• Maintains flexibility with destination-first logic

OpenRouteService Setup:
1. Sign up for free API key at: https://openrouteservice.org/dev/#/signup
2. Replace ORS_API_KEY in code with your actual API key  
3. Restart application to enable road distances
Free tier: 2000 requests per day (sufficient for most use cases)

Equipment weights are stored in equipment_weights.json for easy maintenance and backup."""
        
        tk.Label(about_frame, text=about_text, justify="left", font=('Arial', 8)).pack(padx=10, pady=10)

    def init_optimizer_state(self):
        """Initialize optimizer state variables"""
        self.file_path = None
        self.df = None
        self.date_columns = []
        self.selected_date_columns = []
        self.equipment_qty_vars = {}
        self.equipment_weight_vars = {}
        self.ordered_codes = []
        self.last_results = None
        self.processing = False
        self.date_info_mapping = {}
        self.enhanced_date_info = {}
        self.column_mapping = {}

    def load_saved_settings(self):
        """Load saved settings from config including regional settings"""
        if 'destination' in self.config:
            self.destination_var.set(self.config['destination'])
        else:
            self.destination_dropdown.current(list(WAREHOUSE_COORDS.keys()).index("NEWYORK"))
        
        # Load preferred source setting
        if 'preferred_source' in self.config:
            preferred = self.config['preferred_source']
            if preferred and preferred in WAREHOUSE_COORDS:
                self.preferred_source_var.set(preferred)
        
        # NEW: Load regional filter setting
        if 'region_filter' in self.config:
            region = self.config['region_filter']
            if region and region in REGIONS:
                self.region_var.set(region)

    def save_settings(self):
        """Save current settings to config including regional settings"""
        preferred_source = self.preferred_source_var.get()
        if preferred_source == "(None - Use Standard)":
            preferred_source = None
        
        # NEW: Handle regional filter
        region_filter = self.region_var.get()
        if region_filter == "(None - Use All Warehouses)":
            region_filter = None
        
        self.config.update({
            'destination': self.destination_var.get(),
            'preferred_source': preferred_source,
            'region_filter': region_filter,
            'selected_date_columns': self.selected_date_columns
        })
        ConfigManager.save_config(self.config)

    # Weight Management Methods
    def refresh_weight_list(self):
        """Refresh the weight list display"""
        # Clear existing items
        for item in self.weight_tree.get_children():
            self.weight_tree.delete(item)
        
        # Get all equipment codes (from loaded inventory if available)
        all_equipment = set(self.weight_manager.weights.keys())
        if hasattr(self, 'ordered_codes') and self.ordered_codes:
            all_equipment.update(self.ordered_codes)
        
        # Filter by search term if any
        search_term = self.weight_search_var.get().upper()
        
        for equipment in sorted(all_equipment):
            if search_term and search_term not in equipment.upper():
                continue
                
            weight = self.weight_manager.get_weight(equipment)
            status = "✓ Set" if weight > 0 else "⚠ Not Set"
            
            self.weight_tree.insert('', 'end', values=(equipment, f"{weight:.1f}", status))
    
    def filter_weight_list(self, *args):
        """Filter weight list based on search term"""
        self.refresh_weight_list()
    
    def edit_weight(self, event):
        """Edit weight for selected equipment"""
        selection = self.weight_tree.selection()
        if selection:
            item = self.weight_tree.item(selection[0])
            equipment = item['values'][0]
            weight = self.weight_manager.get_weight(equipment)
            
            self.edit_equipment_var.set(equipment)
            self.edit_weight_var.set(str(weight) if weight > 0 else "")
    
    def save_equipment_weight(self):
        """Save equipment weight"""
        equipment = self.edit_equipment_var.get().strip().upper()
        weight_str = self.edit_weight_var.get().strip()
        
        if not equipment:
            messagebox.showerror("Error", "Please enter an equipment code")
            return
        
        try:
            weight = float(weight_str) if weight_str else 0.0
            if weight < 0:
                messagebox.showerror("Error", "Weight cannot be negative")
                return
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid weight")
            return
        
        # Save weight
        self.weight_manager.set_weight(equipment, weight)
        
        # Update display
        self.refresh_weight_list()
        
        # Update equipment list if it exists
        if hasattr(self, 'equipment_qty_vars') and self.equipment_qty_vars:
            self.update_equipment_list_weights()
        
        # Clear edit fields
        self.edit_equipment_var.set("")
        self.edit_weight_var.set("")
        
        self.weight_status_label.config(text=f"✅ Saved weight for {equipment}: {weight} lbs", foreground="green")
    
    def delete_equipment_weight(self):
        """Delete equipment weight"""
        equipment = self.edit_equipment_var.get().strip().upper()
        
        if not equipment:
            messagebox.showerror("Error", "Please select an equipment code to delete")
            return
        
        if equipment not in self.weight_manager.weights:
            messagebox.showerror("Error", f"No weight found for {equipment}")
            return
        
        # Confirm deletion
        if messagebox.askyesno("Confirm Delete", f"Delete weight for {equipment}?"):
            del self.weight_manager.weights[equipment]
            self.weight_manager.save_weights()
            self.refresh_weight_list()
            
            # Update equipment list if it exists
            if hasattr(self, 'equipment_qty_vars') and self.equipment_qty_vars:
                self.update_equipment_list_weights()
            
            # Clear edit fields
            self.edit_equipment_var.set("")
            self.edit_weight_var.set("")
            
            self.weight_status_label.config(text=f"✅ Deleted weight for {equipment}", foreground="green")
    
    def load_weights_from_csv(self):
        """Load weights from CSV file"""
        file_path = filedialog.askopenfilename(
            title="Select CSV file with equipment weights",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not file_path:
            return
        
        try:
            df = pd.read_csv(file_path)
            
            # Try to find equipment and weight columns
            equipment_col = None
            weight_col = None
            
            for col in df.columns:
                col_lower = str(col).lower()
                if 'equipment' in col_lower or 'code' in col_lower:
                    equipment_col = col
                elif 'weight' in col_lower or 'lbs' in col_lower or 'pounds' in col_lower:
                    weight_col = col
            
            if equipment_col is None or weight_col is None:
                # Let user select columns
                col_dialog = tk.Toplevel(self.root)
                col_dialog.title("Select Columns")
                col_dialog.geometry("400x200")
                col_dialog.transient(self.root)
                col_dialog.grab_set()
                
                tk.Label(col_dialog, text="Select the columns:", font=("Arial", 12, "bold")).pack(pady=10)
                
                tk.Label(col_dialog, text="Equipment Column:").pack()
                equipment_var = tk.StringVar(value=df.columns[0])
                ttk.Combobox(col_dialog, textvariable=equipment_var, values=list(df.columns), width=30).pack(pady=5)
                
                tk.Label(col_dialog, text="Weight Column:").pack()
                weight_var = tk.StringVar(value=df.columns[1] if len(df.columns) > 1 else df.columns[0])
                ttk.Combobox(col_dialog, textvariable=weight_var, values=list(df.columns), width=30).pack(pady=5)
                
                result = {'confirmed': False}
                
                def confirm():
                    result['equipment_col'] = equipment_var.get()
                    result['weight_col'] = weight_var.get()
                    result['confirmed'] = True
                    col_dialog.destroy()
                
                ttk.Button(col_dialog, text="OK", command=confirm).pack(pady=10)
                ttk.Button(col_dialog, text="Cancel", command=col_dialog.destroy).pack()
                
                col_dialog.wait_window()
                
                if not result.get('confirmed'):
                    return
                    
                equipment_col = result['equipment_col']
                weight_col = result['weight_col']
            
            # Load weights
            loaded_count = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    equipment = str(row[equipment_col]).strip().upper()
                    weight = float(row[weight_col])
                    
                    if equipment and weight >= 0:
                        self.weight_manager.weights[equipment] = weight
                        loaded_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {idx + 1}: {str(e)}")
            
            # Save weights
            self.weight_manager.save_weights()
            self.refresh_weight_list()
            
            # Update equipment list if it exists
            if hasattr(self, 'equipment_qty_vars') and self.equipment_qty_vars:
                self.update_equipment_list_weights()
            
            # Show result
            message = f"Loaded {loaded_count} equipment weights from CSV"
            if errors:
                message += f"\n{len(errors)} errors occurred"
                if len(errors) <= 5:
                    message += ":\n" + "\n".join(errors)
            
            messagebox.showinfo("CSV Import Complete", message)
            self.weight_status_label.config(text=f"✅ Loaded {loaded_count} weights from CSV", foreground="green")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load CSV: {str(e)}")
            self.weight_status_label.config(text="❌ CSV load failed", foreground="red")
    
    def export_weights_to_csv(self):
        """Export weights to CSV file"""
        if not self.weight_manager.weights:
            messagebox.showinfo("No Data", "No equipment weights to export")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Save weights as CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not file_path:
            return
        
        try:
            # Create DataFrame
            data = []
            for equipment, weight in sorted(self.weight_manager.weights.items()):
                data.append({
                    'Equipment_Code': equipment,
                    'Weight_lbs': weight
                })
            
            df = pd.DataFrame(data)
            df.to_csv(file_path, index=False)
            
            messagebox.showinfo("Export Complete", f"Exported {len(data)} equipment weights to CSV")
            self.weight_status_label.config(text=f"✅ Exported {len(data)} weights to CSV", foreground="green")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export CSV: {str(e)}")
            self.weight_status_label.config(text="❌ CSV export failed", foreground="red")
    
    def update_equipment_list_weights(self):
        """Update weight display in equipment list"""
        if hasattr(self, 'equipment_qty_vars') and self.equipment_qty_vars:
            # Recreate equipment list with current weights
            self.populate_equipment_list(self.ordered_codes)

    # Quote Converter Methods
    def browse_input_file(self):
        filename = filedialog.askopenfilename(
            title="Select Input Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.input_file.set(filename)
            if not self.output_file.get():
                base_name = Path(filename).stem
                output_name = f"{base_name}_converted.xlsx"
                output_path = Path(filename).parent / output_name
                self.output_file.set(str(output_path))
    
    def browse_output_file(self):
        filename = filedialog.asksaveasfilename(
            title="Save Converted File As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.output_file.set(filename)
    
    def log_quote_message(self, message):
        """Add message to quote results text area"""
        self.quote_results_text.insert(tk.END, message + "\n")
        self.quote_results_text.see(tk.END)
        self.root.update()
    
    def convert_quote(self):
        # Validate inputs
        if not self.input_file.get():
            messagebox.showerror("Error", "Please select an input file")
            return
        
        if not self.output_file.get():
            messagebox.showerror("Error", "Please specify an output file")
            return
        
        if not os.path.exists(self.input_file.get()):
            messagebox.showerror("Error", "Input file does not exist")
            return
        
        try:
            self.quote_progress.start()
            self.quote_status_label.config(text="Converting...", foreground="orange")
            self.quote_results_text.delete(1.0, tk.END)
            
            self.log_quote_message("Starting conversion...")
            self.log_quote_message(f"Input file: {self.input_file.get()}")
            self.log_quote_message(f"Output file: {self.output_file.get()}")
            
            # Read the input Excel file
            self.log_quote_message("Reading input Excel file...")
            
            df_raw = pd.read_excel(self.input_file.get(), sheet_name=0, header=None)
            self.log_quote_message(f"Raw file contains {len(df_raw)} rows and {len(df_raw.columns)} columns")
            
            # Look for the header row
            header_row = None
            for i in range(min(5, len(df_raw))):
                row_values = df_raw.iloc[i].astype(str).str.lower()
                if any('main' in str(val).lower() for val in df_raw.iloc[i]) and \
                   any('equipment' in str(val).lower() for val in df_raw.iloc[i]) and \
                   any('description' in str(val).lower() for val in df_raw.iloc[i]):
                    header_row = i
                    self.log_quote_message(f"Found header row at index {i}")
                    break
            
            if header_row is None:
                header_row = 1
                self.log_quote_message(f"Header row not auto-detected, using row {header_row + 1}")
            
            # Read the file with the correct header row
            df = pd.read_excel(self.input_file.get(), sheet_name=0, header=header_row)
            
            self.log_quote_message(f"Data contains {len(df)} rows and {len(df.columns)} columns")
            self.log_quote_message(f"Column names: {list(df.columns)[:10]}...")
            
            # Check if required columns exist
            required_columns = ['Main', 'Equipment', 'Description', 'Ordered']
            optional_columns = ['Qty Available', 'Min Avail Qty']  # Both needed for partial availability
            column_mapping = {}
            missing_columns = []

            for req_col in required_columns:
                found = False
                for actual_col in df.columns:
                    if str(actual_col).strip().lower() == req_col.lower():
                        column_mapping[req_col] = actual_col
                        found = True
                        break
                if not found:
                    missing_columns.append(req_col)

            # Check for optional columns (needed for partial availability detection)
            for opt_col in optional_columns:
                found = False
                for actual_col in df.columns:
                    col_lower = str(actual_col).strip().lower()
                    opt_lower = opt_col.lower()
                    # Match exact or without spaces
                    if col_lower == opt_lower or col_lower.replace(' ', '') == opt_lower.replace(' ', ''):
                        column_mapping[opt_col] = actual_col
                        found = True
                        self.log_quote_message(f"✓ Found optional column: '{actual_col}' -> {opt_col}")
                        break
                if not found:
                    self.log_quote_message(f"⚠️  Optional column '{opt_col}' not found - partial availability detection may not work properly")
            
            if missing_columns:
                self.log_quote_message("Available columns in file:")
                for col in df.columns:
                    self.log_quote_message(f"  - '{col}'")
                error_msg = f"Missing required columns: {missing_columns}"
                self.log_quote_message(f"ERROR: {error_msg}")
                messagebox.showerror("Error", error_msg)
                return
            
            self.log_quote_message(f"Column mapping: {column_mapping}")

            # Extract required and optional columns
            self.log_quote_message("Extracting columns...")
            converted_df = pd.DataFrame()

            # Extract required columns
            for req_col in required_columns:
                actual_col = column_mapping[req_col]
                converted_df[req_col] = df[actual_col]

            # Extract optional columns if found
            for opt_col in optional_columns:
                if opt_col in column_mapping:
                    actual_col = column_mapping[opt_col]
                    converted_df[opt_col] = df[actual_col]
                    self.log_quote_message(f"  ✓ Including optional column: {opt_col}")
            
            # Clean the data
            converted_df = converted_df.dropna(how='all')
            converted_df = converted_df[
                ~(converted_df['Main'].isna() & 
                  converted_df['Equipment'].isna() & 
                  converted_df['Description'].isna())
            ]
            
            self.log_quote_message(f"Filtered data contains {len(converted_df)} rows")
            converted_df = converted_df.fillna('')
            
            # Save to output file
            self.log_quote_message("Saving converted file...")
            with pd.ExcelWriter(self.output_file.get(), engine='openpyxl') as writer:
                converted_df.to_excel(writer, sheet_name='Rental Items', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Rental Items']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Show conversion summary
            self.log_quote_message("=" * 50)
            self.log_quote_message("CONVERSION COMPLETED SUCCESSFULLY!")
            self.log_quote_message("=" * 50)
            self.log_quote_message(f"Original columns: {len(df.columns)}")
            self.log_quote_message(f"Converted columns: {len(converted_df.columns)}")
            self.log_quote_message(f"Rows processed: {len(converted_df)}")
            self.log_quote_message(f"Output saved to: {self.output_file.get()}")
            
            # Show sample of converted data
            self.log_quote_message("\nSample of converted data:")
            self.log_quote_message("-" * 30)
            for i, row in converted_df.head(5).iterrows():
                self.log_quote_message(f"Row {i+1}: {row['Main']} | {row['Equipment']} | {row['Description'][:30]}... | {row['Ordered']}")
            
            self.quote_status_label.config(text="Conversion completed successfully!", foreground="green")
            messagebox.showinfo("Success", "Quote format converted successfully!")
            
        except Exception as e:
            error_msg = f"Error during conversion: {str(e)}"
            self.log_quote_message(f"ERROR: {error_msg}")
            self.quote_status_label.config(text="Conversion failed", foreground="red")
            messagebox.showerror("Error", error_msg)
        
        finally:
            self.quote_progress.stop()

    # File loading methods
    def load_inventory_file(self):
        """Load inventory Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select Inventory Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_path:
            return
        
        self.opt_progress.start()
        
        def load_in_thread():
            try:
                print("DEBUG: Loading formatted Excel file...")
                
                wb = load_workbook(filename=file_path, read_only=True)
                ws = wb.active
                
                print("DEBUG: Reading file structure...")
                row1_values = [cell.value for cell in ws[1]]
                row2_values = [cell.value for cell in ws[2]]
                row3_values = [cell.value for cell in ws[3]]
                
                print(f"DEBUG: Row 1 (Title): {row1_values[:10]}...")
                print(f"DEBUG: Row 2 (Dates): {row2_values[:10]}...")
                print(f"DEBUG: Row 3 (Headers): {row3_values[:10]}...")
                
                header_row_idx = 3
                equipment_col = None
                location_col = None
                
                for i, header in enumerate(row3_values):
                    if header and str(header).upper().strip() == 'EQUIPMENT':
                        equipment_col = i
                    elif header and str(header).upper().strip() == 'LOCATION':
                        location_col = i
                
                if equipment_col is None or location_col is None:
                    raise Exception("Could not find Equipment and Location columns in header row")
                
                print(f"DEBUG: Equipment column at index {equipment_col}, Location at index {location_col}")
                
                df = pd.read_excel(file_path, header=header_row_idx-1)
                print(f"DEBUG: Loaded {len(df)} rows, {len(df.columns)} columns")
                print(f"DEBUG: Column names: {list(df.columns)}")
                
                df = df[df['Equipment'].notna()].copy()
                df = df[df['Equipment'] != ''].copy()
                print(f"DEBUG: After cleaning: {len(df)} rows")
                
                # Identify date columns
                date_columns = []
                date_info_mapping = {}
                
                print("DEBUG: Identifying date columns...")
                for i, (header, date_info) in enumerate(zip(row3_values, row2_values)):
                    if header and str(header).strip() == 'Qty':
                        if date_info and str(date_info).strip():
                            date_str = str(date_info).strip()
                            if i < len(df.columns):
                                actual_col_name = df.columns[i]
                                date_columns.append(actual_col_name)
                                date_info_mapping[actual_col_name] = date_str
                                print(f"DEBUG: Found date column {i}: '{actual_col_name}' with date '{date_str}'")
                
                print(f"DEBUG: Identified {len(date_columns)} date columns: {date_columns}")
                print(f"DEBUG: Date info mapping: {date_info_mapping}")
                
                wb.close()
                
                self.root.after(0, self._finish_file_load_enhanced, file_path, df, date_columns, date_info_mapping)
                
            except Exception as e:
                error_msg = str(e)
                print(f"DEBUG: Error in load_in_thread: {error_msg}")
                import traceback
                traceback.print_exc()
                self.root.after(0, lambda msg=error_msg: messagebox.showerror("Error", f"Failed to load file: {msg}"))
            finally:
                self.root.after(0, self.opt_progress.stop)
        
        threading.Thread(target=load_in_thread, daemon=True).start()
    
    def _finish_file_load_enhanced(self, file_path, df, date_columns, date_info_mapping):
        """Finish file loading with enhanced date column identification"""
        self.df = df
        self.file_path = file_path
        self.file_label.config(text=f"✅ Loaded: {file_path.split('/')[-1]}")
        
        # Store the original date info mapping and initialize column mapping
        self.date_info_mapping = date_info_mapping
        new_column_mapping = {}  # Initialize empty mapping
        self.column_mapping = {}  # Store the column mapping
        
        # Create better column names with date information
        updated_date_columns = []
        enhanced_date_info = {}
        
        print(f"DEBUG: Processing date columns with mapping: {date_info_mapping}")
        
        # Only process if we have date columns
        if date_columns:
            for i, col in enumerate(date_columns):
                if col in date_info_mapping and date_info_mapping[col]:
                    date_info = str(date_info_mapping[col]).strip()
                    
                    # Create more descriptive column names
                    if 'today' in date_info.lower():
                        print(f"DEBUG: Skipping 'Today' column: {col}")
                        continue
                        
                    elif any(day in date_info for day in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']):
                        # Parse day and date information
                        clean_date = date_info.replace('+', '').strip()
                        new_name = f"Date_{i+1}_{clean_date.replace(' ', '_').replace('/', '_')}"
                        enhanced_info = f"Available {clean_date}"
                    elif any(month in date_info for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                                                            'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']):
                        clean_date = date_info.replace('+', '').strip()
                        new_name = f"Date_{i+1}_{clean_date.replace(' ', '_').replace('/', '_')}"
                        enhanced_info = f"Available {clean_date}"
                    else:
                        # Generic date info
                        clean_info = date_info.replace(' ', '_').replace('+', '').replace('/', '_')
                        new_name = f"Date_{i+1}_{clean_info}"
                        enhanced_info = f"Available {date_info}"
                    
                    new_column_mapping[col] = new_name
                    updated_date_columns.append(new_name)
                    enhanced_date_info[new_name] = enhanced_info
                    print(f"DEBUG: Mapping {col} -> {new_name} ({enhanced_info})")
                else:
                    # No date info available, create generic name
                    new_name = f"Date_Column_{i+1}"
                    updated_date_columns.append(new_name)
                    enhanced_date_info[new_name] = f"Date Column {i+1}"
                    new_column_mapping[col] = new_name
                    print(f"DEBUG: Generic mapping {col} -> {new_name}")
        
        # Rename columns in dataframe if we have mappings
        if new_column_mapping:
            try:
                self.df = self.df.rename(columns=new_column_mapping)
                self.column_mapping = new_column_mapping  # Store successful mapping
                print(f"DEBUG: Renamed columns in dataframe: {new_column_mapping}")
            except Exception as e:
                print(f"DEBUG: Error renaming columns: {e}")
                # If renaming fails, use original column names
                updated_date_columns = date_columns
                enhanced_date_info = {col: f"Date Column {col}" for col in date_columns}
        else:
            # No mapping needed, use original columns
            updated_date_columns = date_columns if date_columns else []
            enhanced_date_info = {col: f"Date Column {col}" for col in updated_date_columns}
        
        # Store enhanced information for the date selection dialog
        self.date_columns = updated_date_columns
        self.enhanced_date_info = enhanced_date_info
        self.selected_date_columns = updated_date_columns.copy()
        
        print(f"DEBUG: Final date_columns: {self.date_columns}")
        print(f"DEBUG: Enhanced date info: {enhanced_date_info}")
        
        # Create display text for the label
        if len(updated_date_columns) > 0:
            # Show enhanced date information in the display
            if len(updated_date_columns) <= 2:
                display_items = []
                for col in updated_date_columns:
                    if col in enhanced_date_info:
                        display_items.append(enhanced_date_info[col])
                    else:
                        display_items.append(col)
                display_text = ", ".join(display_items)
            else:
                first_info = enhanced_date_info.get(updated_date_columns[0], updated_date_columns[0])
                last_info = enhanced_date_info.get(updated_date_columns[-1], updated_date_columns[-1])
                display_text = f"{first_info} ... {last_info} ({len(updated_date_columns)} total)"
            
            self.dates_label.config(text=f"Auto-selected: {display_text}")
        else:
            self.dates_label.config(text="No date columns found")
        
        # Load equipment codes
        try:
            equipment_codes = list(dict.fromkeys(self.df['Equipment'].dropna()))
            self.ordered_codes = equipment_codes
            self.populate_equipment_list(equipment_codes)
            print(f"DEBUG: Found {len(equipment_codes)} equipment codes")
        except Exception as e:
            print(f"DEBUG: Error loading equipment codes: {e}")
            self.ordered_codes = []
            messagebox.showerror("Error", f"Failed to load equipment codes: {e}")

    def load_needs_excel(self):
        """Load quote Excel file for prefilling quantities"""
        if not self.ordered_codes:
            messagebox.showerror("Error", "Please load an inventory file first to see available equipment codes!")
            return
            
        file_path = filedialog.askopenfilename(
            title="Select Quote Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_path:
            return
        
        try:
            self.opt_progress.start()
            self.file_label.config(text="📋 Loading quote file...")
            
            # Try to read the Excel file
            print(f"DEBUG: Loading quote file: {file_path}")
            
            # First, try to detect the correct sheet and header row
            xl_file = pd.ExcelFile(file_path)
            sheet_names = xl_file.sheet_names
            print(f"DEBUG: Available sheets: {sheet_names}")
            
            # Try the first sheet or look for 'Rental Items' sheet
            sheet_to_use = sheet_names[0]
            if 'Rental Items' in sheet_names:
                sheet_to_use = 'Rental Items'
            elif any('rental' in name.lower() for name in sheet_names):
                sheet_to_use = next(name for name in sheet_names if 'rental' in name.lower())
            
            print(f"DEBUG: Using sheet: {sheet_to_use}")
            
            # Read the file and detect header row
            df_raw = pd.read_excel(file_path, sheet_name=sheet_to_use, header=None)
            print(f"DEBUG: Raw data shape: {df_raw.shape}")
            
            # Look for header row containing Main, Equipment, Description, Ordered
            header_row = None
            for i in range(min(5, len(df_raw))):
                row_values = [str(val).lower() if pd.notnull(val) else '' for val in df_raw.iloc[i]]
                has_main = any('main' in val for val in row_values)
                has_equipment = any('equipment' in val for val in row_values)
                has_ordered = any('ordered' in val or 'qty' in val or 'quantity' in val for val in row_values)
                
                if has_main and has_equipment and has_ordered:
                    header_row = i
                    print(f"DEBUG: Found header row at index {i}")
                    break
            
            if header_row is None:
                # Fallback - try common header positions
                header_row = 0
                print("DEBUG: Using default header row (0)")
            
            # Read with proper header
            df_needs = pd.read_excel(file_path, sheet_name=sheet_to_use, header=header_row)
            print(f"DEBUG: Loaded data shape: {df_needs.shape}")
            print(f"DEBUG: Columns: {list(df_needs.columns)}")
            
            # Find the relevant columns
            equipment_col = None
            quantity_col = None
            
            # Look for equipment column
            for col in df_needs.columns:
                col_name = str(col).lower().strip()
                if 'equipment' in col_name and equipment_col is None:
                    equipment_col = col
                    break
            
            # Look for quantity/ordered column
            for col in df_needs.columns:
                col_name = str(col).lower().strip()
                if any(keyword in col_name for keyword in ['ordered', 'qty', 'quantity', 'need', 'request']):
                    quantity_col = col
                    break
            
            if equipment_col is None:
                # Fallback to column positions (assuming Main=0, Equipment=1, Description=2, Ordered=3)
                if len(df_needs.columns) >= 4:
                    equipment_col = df_needs.columns[1]  # Second column typically equipment
                    quantity_col = df_needs.columns[3]   # Fourth column typically ordered
                    print(f"DEBUG: Using positional columns - Equipment: {equipment_col}, Quantity: {quantity_col}")
                else:
                    raise Exception(f"Could not identify equipment and quantity columns. Available columns: {list(df_needs.columns)}")
            
            if quantity_col is None:
                quantity_col = df_needs.columns[-1]  # Last column as fallback
            
            print(f"DEBUG: Using Equipment column: '{equipment_col}', Quantity column: '{quantity_col}'")
            
            # Extract needs dictionary
            needs_dict = {}
            processed_rows = 0
            
            for idx, row in df_needs.iterrows():
                try:
                    # Get equipment code
                    equipment_raw = row[equipment_col] if pd.notnull(row[equipment_col]) else ''
                    equipment_code = str(equipment_raw).strip().upper()
                    
                    # Get quantity
                    qty_raw = row[quantity_col] if pd.notnull(row[quantity_col]) else 0
                    
                    # Handle different quantity formats
                    if isinstance(qty_raw, str):
                        # Remove non-numeric characters except decimal point
                        qty_clean = ''.join(c for c in qty_raw if c.isdigit() or c == '.')
                        try:
                            qty = float(qty_clean) if qty_clean else 0
                        except:
                            qty = 0
                    else:
                        qty = float(qty_raw) if pd.notnull(qty_raw) else 0
                    
                    qty = int(qty)  # Convert to integer
                    
                    # Only process valid entries
                    if equipment_code and qty > 0:
                        if equipment_code in needs_dict:
                            needs_dict[equipment_code] += qty
                        else:
                            needs_dict[equipment_code] = qty
                        processed_rows += 1
                        print(f"DEBUG: {equipment_code} -> {qty}")
                
                except Exception as e:
                    print(f"DEBUG: Error processing row {idx}: {e}")
                    continue
            
            print(f"DEBUG: Processed {processed_rows} rows, found {len(needs_dict)} unique equipment codes")
            
            # Map to available equipment codes
            prefills = {}
            matches_found = 0
            
            # Create normalized mapping for better matching
            norm_map = {code.strip().upper(): code for code in self.ordered_codes}
            
            for quote_code, qty in needs_dict.items():
                if quote_code in norm_map:
                    actual_code = norm_map[quote_code]
                    prefills[actual_code] = str(qty)
                    matches_found += 1
                    print(f"DEBUG: Matched {quote_code} -> {actual_code} (qty: {qty})")
                else:
                    print(f"DEBUG: No match found for {quote_code}")
            
            # Update the equipment list with prefilled quantities
            self.populate_equipment_list(self.ordered_codes, prefills)
            
            # Update file label with results
            quote_filename = file_path.split('/')[-1].split('\\')[-1]
            self.file_label.config(
                text=f"📋 Quote loaded: {quote_filename} ({matches_found}/{len(needs_dict)} codes matched)"
            )
            
            # Show summary dialog
            summary_msg = f"""Quote File Loaded Successfully!

File: {quote_filename}
Sheet: {sheet_to_use}

Results:
• Equipment codes in quote: {len(needs_dict)}
• Codes matched to inventory: {matches_found}
• Total items requested: {sum(needs_dict.values())}

{matches_found} equipment quantities have been prefilled.
Unmatched codes were not found in your inventory file."""
            
            messagebox.showinfo("Quote Loaded", summary_msg)
            
            if matches_found == 0:
                messagebox.showwarning(
                    "No Matches", 
                    "No equipment codes from the quote matched your inventory file.\n\n"
                    "This might happen if:\n"
                    "• The quote uses different naming conventions\n"
                    "• The equipment codes don't exist in your inventory\n"
                    "• The wrong columns were detected\n\n"
                    "You can still manually enter quantities."
                )
        
        except Exception as e:
            error_msg = f"Failed to load quote file: {str(e)}"
            print(f"DEBUG: {error_msg}")
            messagebox.showerror("Error", error_msg)
            self.file_label.config(text="❌ Quote load failed")
        
        finally:
            self.opt_progress.stop()

    def populate_equipment_list(self, equipment_codes, prefills=None):
        """Populate the equipment list with quantity inputs and weight display"""
        for widget in self.eq_list_frame.winfo_children():
            widget.destroy()
        
        self.equipment_qty_vars.clear()
        self.equipment_weight_vars.clear()
        
        # Headers
        ttk.Label(self.eq_list_frame, text="Equipment Code", font=("Arial", 10, "bold")).grid(
            row=0, column=0, sticky='w', padx=5, pady=2)
        ttk.Label(self.eq_list_frame, text="Quantity", font=("Arial", 10, "bold")).grid(
            row=0, column=1, padx=5, pady=2)
        ttk.Label(self.eq_list_frame, text="Weight (lbs)", font=("Arial", 10, "bold")).grid(
            row=0, column=2, padx=5, pady=2)
        ttk.Label(self.eq_list_frame, text="Total Weight", font=("Arial", 10, "bold")).grid(
            row=0, column=3, padx=5, pady=2)
        
        for i, code in enumerate(equipment_codes, 1):
            # Equipment code label
            ttk.Label(self.eq_list_frame, text=code).grid(
                row=i, column=0, sticky='w', padx=5, pady=1)
            
            # Quantity entry
            qty_var = tk.StringVar()
            if prefills and code in prefills:
                qty_var.set(prefills[code])
            
            qty_entry = ttk.Entry(self.eq_list_frame, width=10, textvariable=qty_var)
            qty_entry.grid(row=i, column=1, padx=5, pady=1)
            
            # Weight display
            weight = self.weight_manager.get_weight(code)
            weight_var = tk.StringVar(value=f"{weight:.1f}")
            weight_label = ttk.Label(self.eq_list_frame, textvariable=weight_var, 
                                   foreground="blue" if weight > 0 else "gray")
            weight_label.grid(row=i, column=2, padx=5, pady=1)
            
            # Total weight calculation (updates automatically)
            total_weight_var = tk.StringVar(value="0.0")
            total_weight_label = ttk.Label(self.eq_list_frame, textvariable=total_weight_var, 
                                         foreground="green")
            total_weight_label.grid(row=i, column=3, padx=5, pady=1)
            
            # Update total weight when quantity changes
            def update_total_weight(code=code, qty_var=qty_var, total_var=total_weight_var):
                try:
                    qty = int(qty_var.get()) if qty_var.get() else 0
                    weight = self.weight_manager.get_weight(code)
                    total = qty * weight
                    total_var.set(f"{total:.1f}")
                except:
                    total_var.set("0.0")
            
            qty_var.trace_add('write', lambda *args, update_func=update_total_weight: update_func())
            
            self.equipment_qty_vars[code] = qty_var
            self.equipment_weight_vars[code] = (weight_var, total_weight_var)
    
    def clear_all_quantities(self):
        """Clear all quantity entries"""
        for var in self.equipment_qty_vars.values():
            var.set("")
    
    def set_all_quantities(self, value):
        """Set all quantities to a specific value"""
        for var in self.equipment_qty_vars.values():
            var.set(str(value))

    def select_dates(self):
        """Open date column selection dialog"""
        if self.df is None:
            messagebox.showerror("Error", "Load an Excel file first.")
            return
        
        if not self.date_columns:
            messagebox.showerror("Error", "No date columns found in the loaded file.")
            return
        
        win = tk.Toplevel(self.root)
        win.title("Select Date Columns")
        win.geometry("1200x500")
        win.transient(self.root)
        win.grab_set()
        
        tk.Label(win, text="Select the date columns to use for availability calculations:",
                font=("Arial", 11, "bold")).pack(pady=10)
        
        tk.Label(win, text="✓ Checked columns will be used to determine minimum availability across the date range.",
                font=("Arial", 9), foreground="gray").pack(pady=(0, 10))
        
        # Main content frame
        content_frame = tk.Frame(win)
        content_frame.pack(fill='both', expand=True, padx=10)
        
        # Create canvas and scrollbar for the date columns
        canvas = tk.Canvas(content_frame)
        scrollbar = ttk.Scrollbar(content_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        vars_list = []
        
        # Create header
        header_frame = tk.Frame(scrollable_frame)
        header_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Label(header_frame, text="Date Information", font=("Arial", 10, "bold"), width=25, anchor="w").pack(side="left")
        tk.Label(header_frame, text="Selected", font=("Arial", 10, "bold"), width=10).pack(side="left")
        tk.Label(header_frame, text="Sample Quantities", font=("Arial", 10, "bold"), width=20, anchor="w").pack(side="left")
        tk.Label(header_frame, text="Column Name", font=("Arial", 10, "bold"), width=25, anchor="w").pack(side="left")
        
        ttk.Separator(scrollable_frame, orient='horizontal').pack(fill='x', padx=5, pady=5)
        
        for i, col in enumerate(self.date_columns):
            var = tk.IntVar(value=1 if col in self.selected_date_columns else 0)
            vars_list.append((col, var))
            
            # Get enhanced date information if available
            if hasattr(self, 'enhanced_date_info') and col in self.enhanced_date_info:
                date_display = self.enhanced_date_info[col]
            else:
                # Fallback to original date info mapping
                original_col = None
                if hasattr(self, 'date_info_mapping'):
                    # Find the original column name that maps to this renamed column
                    for orig, renamed in getattr(self, 'column_mapping', {}).items():
                        if renamed == col:
                            original_col = orig
                            break
                
                if original_col and hasattr(self, 'date_info_mapping') and original_col in self.date_info_mapping:
                    date_info = self.date_info_mapping[original_col]
                    date_display = f"Available {date_info}" if date_info else col
                else:
                    date_display = col
            
            # Get sample data from this column
            try:
                if col in self.df.columns:
                    sample_data = self.df[col].dropna().head(3).tolist()
                    # Format numbers nicely
                    formatted_samples = []
                    for x in sample_data[:3]:
                        if isinstance(x, (int, float)):
                            if x == int(x):  # Whole number
                                formatted_samples.append(str(int(x)))
                            else:
                                formatted_samples.append(f"{x:.1f}")
                        else:
                            formatted_samples.append(str(x)[:8])
                    
                    sample_text = ", ".join(formatted_samples)
                    if len(sample_data) > 3:
                        sample_text += "..."
                    if not sample_text:
                        sample_text = "(empty column)"
                else:
                    sample_text = "(column not found)"
            except Exception as e:
                sample_text = f"(error: {str(e)[:20]})"
            
            # Create row frame
            row_frame = tk.Frame(scrollable_frame)
            row_frame.pack(fill='x', padx=5, pady=2)
            
            # Date information (most important, shown first)
            date_label = tk.Label(row_frame, text=date_display[:35] + ("..." if len(date_display) > 35 else ""), 
                                 width=25, anchor="w", font=("Arial", 9))
            date_label.pack(side="left")
            
            # Checkbox
            cb_frame = tk.Frame(row_frame)
            cb_frame.pack(side="left", padx=15)
            cb = ttk.Checkbutton(cb_frame, variable=var)
            cb.pack()
            
            # Sample quantities
            sample_label = tk.Label(row_frame, text=sample_text, width=20, anchor="w", 
                                   font=("Arial", 8), foreground="gray")
            sample_label.pack(side="left", padx=10)
            
            # Technical column name (smaller, less important)
            col_label = tk.Label(row_frame, text=col[:25] + ("..." if len(col) > 25 else ""), 
                                width=25, anchor="w", font=("Arial", 8), foreground="darkgray")
            col_label.pack(side="left", padx=5)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Button frame
        btn_frame = tk.Frame(win)
        btn_frame.pack(fill='x', pady=10)
        
        # Selection summary
        summary_label = tk.Label(btn_frame, text="", font=("Arial", 9), foreground="blue")
        summary_label.pack(pady=5)
        
        def update_summary():
            selected_count = sum(var.get() for _, var in vars_list)
            summary_label.config(text=f"{selected_count} of {len(vars_list)} columns selected")
        
        def select_all():
            for _, var in vars_list:
                var.set(1)
            update_summary()
        
        def select_none():
            for _, var in vars_list:
                var.set(0)
            update_summary()
        
        def on_ok():
            selected = [col for col, var in vars_list if var.get()]
            
            if not selected:
                messagebox.showerror("Error", "Please select at least one date column.")
                return
            
            self.selected_date_columns = selected
            
            # Create display text with enhanced date information
            if selected:
                if len(selected) <= 2:
                    display_items = []
                    for col in selected:
                        if hasattr(self, 'enhanced_date_info') and col in self.enhanced_date_info:
                            display_items.append(self.enhanced_date_info[col])
                        else:
                            display_items.append(col)
                    display_text = ", ".join(display_items)
                else:
                    first_col = selected[0]
                    last_col = selected[-1]
                    
                    first_info = first_col
                    last_info = last_col
                    if hasattr(self, 'enhanced_date_info'):
                        first_info = self.enhanced_date_info.get(first_col, first_col)
                        last_info = self.enhanced_date_info.get(last_col, last_col)
                    
                    display_text = f"{first_info} ... {last_info} ({len(selected)} total)"
            else:
                display_text = "No date columns selected"
            
            self.dates_label.config(text=f"Selected: {display_text}")
            
            # Save selection to config
            self.config['selected_date_columns'] = self.selected_date_columns
            ConfigManager.save_config(self.config)
            
            win.destroy()
        
        # Update summary initially
        update_summary()
        
        # Bind checkbox changes to update summary
        for _, var in vars_list:
            var.trace_add('write', lambda *args: update_summary())
        
        # Control buttons
        ttk.Button(btn_frame, text="Select All", command=select_all).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Select None", command=select_none).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="OK", command=on_ok).pack(side="right", padx=5)
        ttk.Button(btn_frame, text="Cancel", command=win.destroy).pack(side="right", padx=5)

    def run_optimizer_threaded(self):
        """Run optimization in separate thread with regional support"""
        if self.processing:
            return
        
        if self.df is None:
            messagebox.showerror("Error", "Please load an Excel file first!")
            return
        
        if not self.selected_date_columns:
            messagebox.showerror("Error", "Please select at least one date column!")
            return
        
        equipment_needs = {}
        for code in self.ordered_codes:
            qty_var = self.equipment_qty_vars[code]
            try:
                qty = int(qty_var.get())
                if qty > 0:
                    equipment_needs[code] = qty
            except ValueError:
                continue
        
        if not equipment_needs:
            messagebox.showerror("Error", "Please enter at least one equipment quantity!")
            return
        
        destination = self.destination_var.get().strip().upper()
        if destination not in WAREHOUSE_COORDS:
            messagebox.showerror("Error", f"Invalid destination: {destination}")
            return
        
        # Get preferred source warehouse
        preferred_source = self.preferred_source_var.get()
        if preferred_source == "(None - Use Standard)" or preferred_source == "":
            preferred_source = None
        
        # NEW: Get regional filter
        region_filter = self.region_var.get()
        if region_filter == "(None - Use All Warehouses)" or region_filter == "":
            region_filter = None
        
        self.processing = True
        self.optimize_button.config(state="disabled", text="⏳ Processing...")
        self.opt_progress.start()
        
        def optimize_in_thread():
            try:
                scenarios = self.engine.optimize_pulls(
                    self.df, equipment_needs, destination, self.selected_date_columns, 
                    preferred_source=preferred_source, region_filter=region_filter
                )
                self.root.after(0, self._finish_optimization, scenarios)
            except Exception as e:
                error_msg = str(e)
                import traceback
                traceback.print_exc()
                self.root.after(0, lambda msg=error_msg: messagebox.showerror("Error", f"Optimization failed: {msg}"))
            finally:
                self.root.after(0, self._finish_processing)
        
        threading.Thread(target=optimize_in_thread, daemon=True).start()

    def _finish_optimization(self, scenarios):
        """Finish optimization on main thread"""
        self.last_results = scenarios
        self.display_results(scenarios)
        self.export_button.config(state="normal")
        self.save_settings()
    
    def _finish_processing(self):
        """Clean up after processing"""
        self.processing = False
        self.optimize_button.config(state="normal", text="🚀 Optimize!")
        self.opt_progress.stop()

    def display_results(self, scenarios: List[OptimizationScenario]):
        """Display optimization results for all scenarios with regional and weight information"""
        self.summary_text.delete("1.0", tk.END)
        self.detail_text.delete("1.0", tk.END)
        
        # Get configuration info for display
        preferred_source = self.preferred_source_var.get()
        if preferred_source == "(None - Use Standard)" or preferred_source == "":
            preferred_source = None
        
        region_filter = self.region_var.get()
        if region_filter == "(None - Use All Warehouses)" or region_filter == "":
            region_filter = None
        
        distance_type = "road-based" if self.engine.use_road_distances else "geodesic"
        
        # Header information
        summary_lines = [
            f"OPTIMIZATION RESULTS WITH REGIONAL SUPPORT ({distance_type.upper()} DISTANCES) - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "=" * 110,
            f"Destination: {self.destination_var.get()}",
            f"Preferred Source: {preferred_source if preferred_source else '(None - Standard distance-based optimization)'}",
            f"Regional Filter: {region_filter if region_filter else '(None - All warehouses available)'}",
            f"Distance Method: {'🛣️ Road-based (OpenRouteService)' if self.engine.use_road_distances else '📏 Geodesic (straight line)'}",
            "",
            "SCENARIO COMPARISON OVERVIEW:",
            "-" * 60
        ]
        
        # Add scenario overview table with weights and regional info
        for i, scenario in enumerate(scenarios, 1):
            total_items = sum(r.requested_quantity for r in scenario.results)
            fulfilled_items = sum(r.requested_quantity - r.shortfall for r in scenario.results)
            
            summary_lines.extend([
                f"{i}. {scenario.name.upper()} SCENARIO:",
                f"   Description: {scenario.description}",
                f"   Fill Rate: {scenario.fill_rate:.1f}% ({fulfilled_items}/{total_items} items)",
                f"   Total Distance: {scenario.total_distance:,.1f} miles ({distance_type})",
                f"   Total Weight: {scenario.total_weight:,.1f} lbs",
                f"   Warehouse Trips: {scenario.total_trips} ({scenario.consolidated_trips} consolidated)",
                ""
            ])
        
        if region_filter:
            summary_lines.extend([
                f"🌍 REGIONAL INFORMATION:",
                f"Selected Region: {region_filter}",
                f"Regional Warehouses: {', '.join(REGIONS[region_filter]['warehouses'])}",
                f"Description: {REGIONS[region_filter]['description']}",
                ""
            ])
        
        # Detailed breakdown for each scenario
        for scenario in scenarios:
            summary_lines.extend([
                f"{'='*25} {scenario.name.upper()} SCENARIO DETAILS {'='*25}",
                f"Strategy: {scenario.description}",
                f"Total Weight: {scenario.total_weight:,.1f} lbs",
                f"Distance Type: {distance_type} distances",
                ""
            ])
            
            # Calculate trip consolidation info for this scenario with regional data
            warehouse_trips = {}
            warehouse_regions = {}
            for r in scenario.results:
                for pull in r.pulls:
                    if pull.warehouse not in warehouse_trips:
                        warehouse_trips[pull.warehouse] = []
                        warehouse_regions[pull.warehouse] = pull.region
                    warehouse_trips[pull.warehouse].append(f"{r.equipment_code}({pull.quantity})")
            
            # Show trip consolidation with weights and regional info
            summary_lines.append("WAREHOUSE TRIPS:")
            for warehouse, items in warehouse_trips.items():
                region_info = warehouse_regions.get(warehouse, "Unknown Region")
                indicator = ""
                if warehouse == self.destination_var.get():
                    indicator = " (LOCAL)"
                elif warehouse == preferred_source:
                    indicator = " ⭐ (PREFERRED SOURCE)"
                
                # Calculate total weight for this warehouse
                warehouse_weight = 0
                for r in scenario.results:
                    for pull in r.pulls:
                        if pull.warehouse == warehouse:
                            warehouse_weight += pull.total_weight
                
                if len(items) > 1:
                    summary_lines.append(f"📦 {warehouse}{indicator} ({region_info}): {', '.join(items)} (CONSOLIDATED) - {warehouse_weight:.1f} lbs")
                else:
                    summary_lines.append(f"📦 {warehouse}{indicator} ({region_info}): {items[0]} - {warehouse_weight:.1f} lbs")
            
            # Show equipment breakdown with weights and regional info
            summary_lines.extend([
                "",
                "EQUIPMENT BREAKDOWN:",
                "-" * 40
            ])
            
            for result in scenario.results:
                summary_lines.append(f"\n{result.equipment_code} - {result.equipment_description}")
                summary_lines.append(f"(Requested: {result.requested_quantity}, Weight per unit: {result.weight_per_unit:.1f} lbs)")
                
                if result.pulls:
                    total_pulled = sum(p.quantity for p in result.pulls)
                    total_weight = sum(p.total_weight for p in result.pulls)
                    summary_lines.append(f"  ✅ PULLS: {total_pulled} units from {len(result.pulls)} location(s) - Total: {total_weight:.1f} lbs")
                    
                    for i, pull in enumerate(result.pulls, 1):
                        indicator = ""
                        if pull.warehouse == self.destination_var.get():
                            indicator = " (LOCAL)"
                        elif pull.warehouse == preferred_source and scenario.name == "Preferred Source":
                            indicator = " ⭐ (PREFERRED)"
                        
                        distance_text = f"{pull.distance} miles" if pull.distance > 0 else "local"
                        summary_lines.append(f"    {i}. {pull.warehouse}{indicator} ({pull.region}): {pull.quantity} units ({distance_text}) - {pull.total_weight:.1f} lbs")
                
                if result.shortfall > 0:
                    shortfall_weight = result.shortfall * result.weight_per_unit
                    summary_lines.append(f"  ❌ Shortfall: {result.shortfall} units ({shortfall_weight:.1f} lbs)")
            
            summary_lines.extend(["", ""])

        # Add Partial Availability Summary
        partial_avail_count = 0
        for scenario in scenarios:
            for result in scenario.results:
                for pull in result.pulls:
                    if pull.is_partial_availability:
                        partial_avail_count += 1

        if partial_avail_count > 0:
            summary_lines.extend([
                "",
                "🔄 PARTIAL AVAILABILITY DETECTED:",
                f"    {partial_avail_count} warehouse(s) have partial availability (Min Avail Qty > Qty Available)",
                "    Equipment is available locally for part of the period, with backfill from other locations",
                "    See detailed view for backfill sources and quantities",
                ""
            ])

        self.summary_text.insert(tk.END, "\n".join(summary_lines))
        
        # Enhanced detailed view showing technical details with regional and weight information
        detail_lines = [
            f"TECHNICAL OPTIMIZATION DETAILS WITH REGIONAL SUPPORT ({distance_type.upper()} DISTANCES) - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "=" * 120,
            f"Destination: {self.destination_var.get()}",
            f"Regional Filter: {region_filter if region_filter else '(All warehouses available)'}",
            f"Date Columns Used: {', '.join(self.selected_date_columns)}",
            f"Distance Method: {'🛣️ Road-based via OpenRouteService' if self.engine.use_road_distances else '📏 Geodesic (straight line)'}",
            ""
        ]
        
        if region_filter:
            detail_lines.extend([
                f"🌍 REGIONAL CONSTRAINTS:",
                f"Region: {region_filter}",
                f"Warehouses: {', '.join(REGIONS[region_filter]['warehouses'])}",
                f"Description: {REGIONS[region_filter]['description']}",
                ""
            ])
        
        for scenario in scenarios:
            detail_lines.extend([
                f"{'='*35} {scenario.name.upper()} SCENARIO {'='*35}",
                f"Strategy: {scenario.description}",
                f"Total Distance: {scenario.total_distance:,.1f} miles ({distance_type})",
                f"Total Weight: {scenario.total_weight:,.1f} lbs",
                f"Fill Rate: {scenario.fill_rate:.1f}%",
                f"Warehouse Trips: {scenario.total_trips} (Consolidated: {scenario.consolidated_trips})",
                "",
                "WAREHOUSE TRIP PLAN WITH REGIONAL INFO:",
                "-" * 50
            ])
            
            # Show trip plan by warehouse with weights and regional info
            warehouse_trips = {}
            warehouse_weights = {}
            warehouse_regions = {}
            for r in scenario.results:
                for pull in r.pulls:
                    if pull.warehouse not in warehouse_trips:
                        warehouse_trips[pull.warehouse] = []
                        warehouse_weights[pull.warehouse] = 0
                        warehouse_regions[pull.warehouse] = pull.region
                    warehouse_trips[pull.warehouse].append(f"{r.equipment_code}({pull.quantity})")
                    warehouse_weights[pull.warehouse] += pull.total_weight
            
            for warehouse, items in sorted(warehouse_trips.items(), key=lambda x: x[0]):
                weight = warehouse_weights[warehouse]
                region = warehouse_regions.get(warehouse, "Unknown Region")
                
                if warehouse == self.destination_var.get():
                    detail_lines.append(f"🏠 {warehouse} (LOCAL - {region}) - {weight:.1f} lbs:")
                elif warehouse == preferred_source and scenario.name == "Preferred Source":
                    detail_lines.append(f"⭐ {warehouse} (PREFERRED SOURCE - {region}) - {weight:.1f} lbs:")
                else:
                    dest_coords = WAREHOUSE_COORDS[self.destination_var.get()]
                    wh_coords = WAREHOUSE_COORDS[warehouse]
                    distance = int(round(self.engine.calculate_distance(wh_coords, dest_coords)))
                    detail_lines.append(f"🚛 {warehouse} ({distance} miles {distance_type} - {region}) - {weight:.1f} lbs:")
                
                for item in items:
                    detail_lines.append(f"   • {item}")
            
            detail_lines.extend(["", "EQUIPMENT PULL DETAILS WITH REGIONAL INFO:", "-" * 50])
            
            for result in scenario.results:
                detail_lines.extend([
                    f"\n{result.equipment_code} - {result.equipment_description}",
                    f"Need: {result.requested_quantity}, Got: {result.requested_quantity - result.shortfall}, Short: {result.shortfall}",
                    f"Weight per unit: {result.weight_per_unit:.1f} lbs, Total weight: {result.total_weight:.1f} lbs"
                ])
                
                if result.pulls:
                    for i, pull in enumerate(result.pulls, 1):
                        detail_lines.append(
                            f"  {i}. {pull.warehouse:<12} | {pull.region:<20} | {pull.quantity:3d} units | "
                            f"{pull.distance:5d} miles | {pull.total_weight:6.1f} lbs | Available: {pull.available_quantity}"
                        )
                else:
                    detail_lines.append("  (No pulls - not available)")
            
            detail_lines.extend(["", ""])

        # Add Partial Availability Details Section
        partial_avail_details = []
        for scenario in scenarios:
            for result in scenario.results:
                for pull in result.pulls:
                    if pull.is_partial_availability:
                        # Find the backfill pull if it exists
                        backfill_pull = next((p for p in result.pulls if p.is_backfill and p.backfill_for_warehouse == pull.warehouse), None)

                        partial_avail_details.append({
                            'scenario': scenario.name,
                            'equipment': result.equipment_code,
                            'description': result.equipment_description,
                            'local_warehouse': pull.warehouse,
                            'region': pull.region,
                            'qty_available': pull.raw_available,
                            'min_avail_qty': pull.min_avail_qty,
                            'backfill_needed': pull.min_avail_qty - pull.raw_available,
                            'backfill_source': backfill_pull.warehouse if backfill_pull else None,
                            'backfill_qty': backfill_pull.quantity if backfill_pull else 0,
                            'backfill_distance': backfill_pull.distance if backfill_pull else 0
                        })

        if partial_avail_details:
            detail_lines.extend([
                "",
                "=" * 120,
                "🔄 PARTIAL AVAILABILITY - BACKFILL REQUIRED",
                "=" * 120,
                "",
                "The following equipment has PARTIAL AVAILABILITY (Min Avail Qty > Qty Available).",
                "Local warehouse can supply units for the full rental period, but additional units must be backfilled.",
                ""
            ])

            for detail in partial_avail_details:
                backfill_status = "✓ BACKFILL ARRANGED" if detail['backfill_source'] else "⚠️  BACKFILL NOT AVAILABLE"
                detail_lines.extend([
                    f"{backfill_status}: {detail['equipment']} - {detail['description']}",
                    f"  • Scenario: {detail['scenario']}",
                    f"  • Local Warehouse: {detail['local_warehouse']} ({detail['region']})",
                    f"  • Qty Available (Full Period): {detail['qty_available']} units",
                    f"  • Min Avail Qty (Total Needed): {detail['min_avail_qty']} units",
                    f"  • Backfill Needed: {detail['backfill_needed']} units",
                    ""
                ])

                if detail['backfill_source']:
                    detail_lines.extend([
                        f"  • Backfill Source: {detail['backfill_source']}",
                        f"  • Backfill Quantity: {detail['backfill_qty']} units",
                        f"  • Backfill Distance: {detail['backfill_distance']} miles",
                        f"  • Notes: {detail['qty_available']} units from {detail['local_warehouse']} + {detail['backfill_qty']} units from {detail['backfill_source']}",
                        ""
                    ])
                else:
                    detail_lines.extend([
                        f"  • WARNING: No backfill source available - {detail['backfill_needed']} units SHORT",
                        ""
                    ])

        self.detail_text.insert(tk.END, "\n".join(detail_lines))

    def sanitize_sheet_name(self, name):
        """Sanitize sheet name by removing invalid characters for Excel"""
        # Excel sheet names cannot contain: [ ] : * ? / \
        invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '-')
        # Excel sheet names must be 31 characters or less
        return sanitized[:31]

    def format_clean_date(self, date_val) -> str:
        """Format a date value (str or datetime) to a clean short date string (M/D/YY)"""
        if pd.isna(date_val):
            return ""
        
        str_val = str(date_val).strip()
        
        # 1. Regex Extraction for Messy Headers (e.g. "Date_8_7WedJan_28")
        # Looks for Month Name followed cleanly by Day Number
        import re
        match = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[^\d]*(\d+)', str_val, re.IGNORECASE)
        if match:
            try:
                month_str = match.group(1)
                day = int(match.group(2))
                
                # Parse month name
                dt = datetime.strptime(month_str, '%b')
                month = dt.month
                year = 26 # User requested year 2026 context
                
                # Verified format: M/D/YY
                return f"{month}/{day}/{year}"
            except:
                pass

        # 2. Try standard parsing if regex failed / not a messy header
        dt_obj = None
        if isinstance(date_val, (datetime, pd.Timestamp)):
            dt_obj = date_val
        else:
            try:
                dt_obj = pd.to_datetime(str_val, errors='coerce')
            except:
                pass
        
        # 3. Format if we have a valid datetime
        if dt_obj is not None and not pd.isna(dt_obj):
            # Requested format: 2/23/26 (M/D/YY)
            return f"{dt_obj.month}/{dt_obj.day}/{dt_obj.strftime('%y')}"
            
        return str_val.split(' ')[0]

    def export_to_excel(self):
        """Export results to Excel with all scenarios including regional and weight information"""
        if not self.last_results:
            messagebox.showinfo("Nothing to Export", "No results to export. Run optimization first.")
            return
        
        distance_type = "Road" if self.engine.use_road_distances else "Geodesic"
        
        file_path = filedialog.asksaveasfilename(
            title=f"Save Optimization Results with Regional Support, {distance_type} Distances & Weights",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not file_path:
            return
        
        try:
            # Get configuration info
            preferred_source = self.preferred_source_var.get()
            if preferred_source == "(None - Use Standard)" or preferred_source == "":
                preferred_source = None
            
            region_filter = self.region_var.get()
            if region_filter == "(None - Use All Warehouses)" or region_filter == "":
                region_filter = None
            
            destination = self.destination_var.get()
            scenarios = self.last_results
            
            # Create Excel file with multiple sheets
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Summary comparison sheet with regional and weight information
                summary_data = []
                summary_data.extend([
                    [f"Optimization Comparison with Regional Support, {distance_type} Distances & Weights", ""],
                    ["Report Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                    ["Destination", destination],
                    ["Preferred Source", preferred_source if preferred_source else "(None - Standard optimization)"],
                    ["Regional Filter", region_filter if region_filter else "(None - All warehouses)"],
                    ["Distance Method", f"{distance_type} distances ({'OpenRouteService road-based' if self.engine.use_road_distances else 'straight-line geodesic'})"],
                    ["Date Columns Used", ", ".join(self.selected_date_columns)],
                    ["", ""],
                ])
                
                if region_filter:
                    summary_data.extend([
                        ["REGIONAL INFORMATION", ""],
                        ["Selected Region", region_filter],
                        ["Regional Warehouses", ", ".join(REGIONS[region_filter]['warehouses'])],
                        ["Region Description", REGIONS[region_filter]['description']],
                        ["", ""],
                    ])
                
                summary_data.extend([
                    ["SCENARIO COMPARISON", ""],
                    ["Scenario", "Fill Rate", f"Total Distance ({distance_type})", "Total Weight (lbs)", "Warehouse Trips", "Consolidated Trips"]
                ])
                
                for scenario in scenarios:
                    summary_data.append([
                        scenario.name,
                        f"{scenario.fill_rate:.1f}%",
                        f"{scenario.total_distance:,.1f} miles",
                        f"{scenario.total_weight:,.1f} lbs",
                        scenario.total_trips,
                        scenario.consolidated_trips
                    ])
                
                summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value', 'Extra1', 'Extra2', 'Extra3', 'Extra4'])
                summary_df.to_excel(writer, sheet_name='Scenario Comparison', index=False)
                
                # Create a sheet for each scenario with regional and weight information
                for scenario in scenarios:
                    sheet_name = self.sanitize_sheet_name(f"{scenario.name} Scenario")
                    export_rows = []
                    
                    # Add scenario results with regional and weight info
                    for result in scenario.results:
                        for pull in result.pulls:
                            # Determine pull type
                            if pull.is_backfill:
                                location_type = f'BACKFILL FOR {pull.backfill_for_warehouse}'
                            elif pull.warehouse == destination:
                                location_type = 'LOCAL'
                            elif pull.warehouse == preferred_source and scenario.name == "Preferred Source":
                                location_type = 'PREFERRED SOURCE'
                            else:
                                location_type = 'STANDARD'

                            export_rows.append({
                                'Equipment': result.equipment_code,
                                'Description': result.equipment_description,
                                'Requested Qty': result.requested_quantity,
                                'Equipment Location': pull.warehouse,
                                'Warehouse Region': pull.region,
                                'Pulled Qty': pull.quantity,
                                'Raw Available': pull.raw_available,
                                'Min Avail Qty': pull.min_avail_qty if pull.min_avail_qty > 0 else '',
                                'Partial Availability': 'YES' if pull.is_partial_availability else '',
                                'Is Backfill': 'YES' if pull.is_backfill else '',
                                'Backfill For': pull.backfill_for_warehouse if pull.is_backfill else '',
                                'Source Available Qty': pull.available_quantity,
                                'Destination Location': destination,
                                f'Distance ({distance_type} miles)': pull.distance,
                                'Weight per Unit (lbs)': pull.weight_per_unit,
                                'Total Weight (lbs)': pull.total_weight,
                                'Location Type': location_type,
                                'Regional Filter Applied': region_filter if region_filter else 'None',
                                'Shortfall': ''
                            })
                        
                        if result.shortfall > 0:
                            shortfall_weight = result.shortfall * result.weight_per_unit
                            export_rows.append({
                                'Equipment': result.equipment_code,
                                'Description': result.equipment_description,
                                'Requested Qty': result.requested_quantity,
                                'Equipment Location': 'SHORTFALL',
                                'Warehouse Region': 'N/A',
                                'Pulled Qty': '',
                                'Source Available Qty': '',
                                'Destination Location': destination,
                                f'Distance ({distance_type} miles)': '',
                                'Weight per Unit (lbs)': result.weight_per_unit,
                                'Total Weight (lbs)': shortfall_weight,
                                'Location Type': '',
                                'Regional Filter Applied': region_filter if region_filter else 'None',
                                'Shortfall': result.shortfall
                            })
                    
                    if export_rows:
                        scenario_df = pd.DataFrame(export_rows)
                        scenario_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Trip comparison sheet with regional and weight information
                trip_comparison_data = []
                for scenario in scenarios:
                    warehouse_trips = {}
                    warehouse_weights = {}
                    warehouse_regions = {}
                    for r in scenario.results:
                        for pull in r.pulls:
                            if pull.warehouse not in warehouse_trips:
                                warehouse_trips[pull.warehouse] = []
                                warehouse_weights[pull.warehouse] = 0
                                warehouse_regions[pull.warehouse] = pull.region
                            warehouse_trips[pull.warehouse].append(f"{r.equipment_code}({pull.quantity})")
                            warehouse_weights[pull.warehouse] += pull.total_weight
                    
                    for warehouse, items in warehouse_trips.items():
                        location_type = 'LOCAL' if warehouse == destination else \
                                      'PREFERRED SOURCE' if warehouse == preferred_source and scenario.name == "Preferred Source" else 'STANDARD'
                        region = warehouse_regions.get(warehouse, "Unknown Region")
                        
                        for item in items:
                            trip_comparison_data.append({
                                'Scenario': scenario.name,
                                'Warehouse': warehouse,
                                'Warehouse_Region': region,
                                'Equipment_Qty': item,
                                'Location_Type': location_type,
                                'Trip_Type': 'CONSOLIDATED' if len(items) > 1 else 'SINGLE',
                                'Warehouse_Weight_lbs': warehouse_weights[warehouse],
                                f'Distance_Type': distance_type,
                                'Regional_Filter_Applied': region_filter if region_filter else 'None'
                            })
                
                if trip_comparison_data:
                    trip_df = pd.DataFrame(trip_comparison_data)
                    trip_df.to_excel(writer, sheet_name='Trip Comparison', index=False)
                
                # Equipment comparison across all scenarios with regional and weight info
                equipment_comparison = []
                
                # Get all equipment codes
                all_equipment = set()
                for scenario in scenarios:
                    for result in scenario.results:
                        all_equipment.add(result.equipment_code)
                
                for equipment_code in sorted(all_equipment):
                    row = {
                        'Equipment': equipment_code,
                        'Weight_per_Unit_lbs': self.weight_manager.get_weight(equipment_code),
                        'Regional_Filter_Applied': region_filter if region_filter else 'None'
                    }
                    
                    for scenario in scenarios:
                        # Find this equipment in the scenario
                        result = next((r for r in scenario.results if r.equipment_code == equipment_code), None)
                        if result:
                            filled = result.requested_quantity - result.shortfall
                            fill_rate = (filled / result.requested_quantity * 100) if result.requested_quantity > 0 else 0
                            
                            # Get regional info for pulls
                            regions_used = list(set([p.region for p in result.pulls]))
                            
                            row[f'{scenario.name}_Requested'] = result.requested_quantity
                            row[f'{scenario.name}_Filled'] = filled
                            row[f'{scenario.name}_Shortfall'] = result.shortfall
                            row[f'{scenario.name}_Fill_Rate'] = f"{fill_rate:.1f}%"
                            row[f'{scenario.name}_Total_Weight_lbs'] = result.total_weight
                            row[f'{scenario.name}_Warehouses'] = ", ".join([p.warehouse for p in result.pulls])
                            row[f'{scenario.name}_Regions_Used'] = ", ".join(regions_used) if regions_used else ""
                        else:
                            row[f'{scenario.name}_Requested'] = 0
                            row[f'{scenario.name}_Filled'] = 0
                            row[f'{scenario.name}_Shortfall'] = 0
                            row[f'{scenario.name}_Fill_Rate'] = "N/A"
                            row[f'{scenario.name}_Total_Weight_lbs'] = 0
                            row[f'{scenario.name}_Warehouses'] = ""
                            row[f'{scenario.name}_Regions_Used'] = ""
                    
                    equipment_comparison.append(row)
                
                if equipment_comparison:
                    eq_comp_df = pd.DataFrame(equipment_comparison)
                    eq_comp_df.to_excel(writer, sheet_name='Equipment Comparison', index=False)

                # Partial Availability Details Sheet
                partial_avail_details = []
                for scenario in scenarios:
                    for result in scenario.results:
                        for pull in result.pulls:
                            if pull.is_partial_availability:
                                # Find the backfill pull if it exists
                                backfill_pull = next((p for p in result.pulls if p.is_backfill and p.backfill_for_warehouse == pull.warehouse), None)

                                partial_avail_details.append({
                                    'Scenario': scenario.name,
                                    'Equipment': result.equipment_code,
                                    'Description': result.equipment_description,
                                    'Local Warehouse': pull.warehouse,
                                    'Region': pull.region,
                                    'Qty Available (Full Period)': pull.raw_available,
                                    'Min Avail Qty (Total Needed)': pull.min_avail_qty,
                                    'Backfill Needed': pull.min_avail_qty - pull.raw_available,
                                    'Backfill Source': backfill_pull.warehouse if backfill_pull else 'NOT AVAILABLE',
                                    'Backfill Qty': backfill_pull.quantity if backfill_pull else 0,
                                    'Backfill Distance (miles)': backfill_pull.distance if backfill_pull else '',
                                    'Status': 'PARTIAL AVAILABILITY - BACKFILL ARRANGED' if backfill_pull else 'PARTIAL AVAILABILITY - BACKFILL NOT AVAILABLE',
                                    'Notes': f'{pull.raw_available} units available for full period, {pull.min_avail_qty - pull.raw_available} units needed as backfill'
                                })

                if partial_avail_details:
                    partial_df = pd.DataFrame(partial_avail_details)
                    partial_df.to_excel(writer, sheet_name='Partial Availability', index=False)

                # Availability & Backfill Report (New Tab)
                avail_report_data = []
                # Clean date headers for display
                clean_date_headers = [self.format_clean_date(d) for d in self.selected_date_columns]
                # Map original headers to clean headers for data retrieval
                header_map = dict(zip(self.selected_date_columns, clean_date_headers))
                
                for scenario in scenarios:
                    for result in scenario.results:
                        local_pull_found = False
                        for pull in result.pulls:
                            if pull.warehouse == destination:
                                local_pull_found = True
                                
                            # Only add if we have timeline data
                            if pull.daily_timeline:
                                row = {
                                    'Scenario': scenario.name,
                                    'Equipment': result.equipment_code,
                                    'Description': result.equipment_description,
                                    'Location': pull.warehouse,
                                    'Type': 'Backfill' if pull.is_backfill else ('Local' if pull.warehouse == destination else 'Transfer'),
                                    'Status': 'Shortage Detected' if pull.shortage_dates else 'Fully Available',
                                    'Max Shortage Qty': pull.max_daily_shortage,
                                    
                                    # Formatted Conflict Info with Clean Dates
                                    'Conflict Dates': ", ".join([f"{self.format_clean_date(r['start'])} - {self.format_clean_date(r['end'])}" for r in pull.shortage_ranges]) if pull.shortage_ranges else "None",
                                    'Transfer Need': ", ".join([f"{r['max_qty']} units ({self.format_clean_date(r['start'])}-{self.format_clean_date(r['end'])})" for r in pull.shortage_ranges]) if pull.shortage_ranges else "None",
                                    
                                    # Legacy field for compatibility
                                    'Shortage Dates List': ", ".join([self.format_clean_date(d) for d in pull.shortage_dates])
                                }
                                
                                # Add daily quantities with clean headers
                                for original_col, clean_col in header_map.items():
                                    avail = pull.daily_timeline.get(original_col, 0)
                                    shortage = pull.daily_shortages.get(original_col, 0)
                                    
                                    if shortage > 0:
                                        row[clean_col] = f"SHORT (-{shortage})"
                                    else:
                                        row[clean_col] = avail
                                        
                                avail_report_data.append(row)
                        
                        # Fix for missing local data (Zero Inventory Scenario)
                        # If the destination wasn't in the pulls (because avail=0), analyze it manually
                        if not local_pull_found:
                            try:
                                # Manually analyze the destination location
                                local_analysis = self.engine.analyze_daily_availability(
                                    self.df, result.equipment_code, destination, 
                                    self.selected_date_columns, result.requested_quantity
                                )
                                
                                if local_analysis:
                                    row = {
                                        'Scenario': scenario.name,
                                        'Equipment': result.equipment_code,
                                        'Description': result.equipment_description,
                                        'Location': destination,
                                        'Type': 'Local (Zero Stock)',
                                        'Status': 'Shortage Detected' if local_analysis['shortage_dates'] else 'No Availability',
                                        'Max Shortage Qty': local_analysis['max_shortage'],
                                        
                                        'Conflict Dates': ", ".join([f"{self.format_clean_date(r['start'])} - {self.format_clean_date(r['end'])}" for r in local_analysis['shortage_ranges']]) if local_analysis['shortage_ranges'] else "None",
                                        'Transfer Need': ", ".join([f"{r['max_qty']} units ({self.format_clean_date(r['start'])}-{self.format_clean_date(r['end'])})" for r in local_analysis['shortage_ranges']]) if local_analysis['shortage_ranges'] else "None",
                                        'Shortage Dates List': ", ".join([self.format_clean_date(d) for d in local_analysis['shortage_dates']])
                                    }
                                    
                                    # Add daily quantities
                                    diff = local_analysis.get('daily_shortages', {})
                                    timeline = local_analysis.get('timeline', {})
                                    
                                    for original_col, clean_col in header_map.items():
                                        avail = timeline.get(original_col, 0)
                                        shortage = diff.get(original_col, 0)
                                        if shortage > 0:
                                            row[clean_col] = f"SHORT (-{shortage})"
                                        else:
                                            row[clean_col] = avail
                                    
                                    avail_report_data.append(row)
                                    
                            except Exception as e:
                                print(f"Error manually analyzing local availability for {result.equipment_code}: {e}")
                
                if avail_report_data:
                    # Reorder columns to put standard info first, then dates
                    base_cols = [
                        'Scenario', 'Equipment', 'Description', 'Location', 
                        'Type', 'Status', 
                        'Conflict Dates', 'Transfer Need',
                        'Max Shortage Qty', 'Shortage Dates List'
                    ]
                    final_cols = base_cols + [col for col in avail_report_data[0].keys() if col not in base_cols]
                    
                    avail_df = pd.DataFrame(avail_report_data)
                    # Ensure columns order
                    avail_df = avail_df[final_cols]
                    avail_df.to_excel(writer, sheet_name='Availability Report', index=False)

            # Apply formatting
            wb = load_workbook(file_path)
            
            # Format all sheets with different colors
            sheet_colors = {
                'Scenario Comparison': "366092",  # Blue
                'Regional (Region 1 (East-South)) S': "E8F5E8",  # Light Green (truncated)
                'Regional (Region 2 (West-Central))': "E8F5E8",  # Light Green
                'Preferred Source Scenario': "FFB347",  # Orange
                'Standard Scenario': "70AD47",  # Green
                'Backup Scenario': "D5A6BD",  # Purple
                'Backup (Region 1 (East-South)) Sc': "D5A6BD",  # Purple (truncated)
                'Backup (Region 2 (West-Central)) ': "D5A6BD",  # Purple (truncated)
                'Trip Comparison': "FFC000",  # Yellow
                'Equipment Comparison': "D5A6BD",  # Purple
                'Partial Availability': "FF8C00"  # Orange for partial availability
            }
            
            for sheet_name in wb.sheetnames:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Header formatting - use partial matching for sanitized names
                    header_color = "366092"  # Default blue
                    for color_key, color_value in sheet_colors.items():
                        if color_key in sheet_name or sheet_name.startswith(color_key[:20]):
                            header_color = color_value
                            break
                    
                    # Special handling for regional and backup scenarios
                    if 'Regional' in sheet_name:
                        header_color = "E8F5E8"  # Light Green
                    elif 'Backup' in sheet_name and ('Region' in sheet_name):
                        header_color = "D5A6BD"  # Purple
                    
                    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                    header_font = Font(color="FFFFFF" if header_color not in ["FFC000", "E8F5E8"] else "000000", bold=True)
                    
                    # Apply header formatting
                    for cell in ws[1]:
                        if cell.value:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center")
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
                    
                    # Special formatting for shortfalls, preferred sources, and regional highlights
                    if 'Scenario' in sheet_name:
                        shortfall_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        preferred_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                        regional_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        
                        for row in ws.iter_rows(min_row=2):
                            if len(row) > 4 and row[4].value == 'SHORTFALL':
                                for cell in row:
                                    cell.fill = shortfall_fill
                            elif len(row) > 11 and str(row[11].value) == 'PREFERRED SOURCE':
                                for cell in row:
                                    cell.fill = preferred_fill
                            elif len(row) > 4 and region_filter:
                                # Check if warehouse is in the selected region
                                warehouse_value = str(row[4].value) if len(row) > 4 and row[4].value else ""
                                if warehouse_value in REGIONS.get(region_filter, {}).get('warehouses', []):
                                    for cell in row:
                                        cell.fill = regional_fill
            
            wb.save(file_path)
            
            # Count scenarios for success message
            scenario_count = len(scenarios)
            scenario_names = ", ".join([s.name for s in scenarios])
            total_weight = sum(s.total_weight for s in scenarios) / len(scenarios)  # Average weight across scenarios
            
            regional_info = ""
            if region_filter:
                regional_info = f"""
Regional Configuration:
• Selected Region: {region_filter}
• Regional Warehouses: {', '.join(REGIONS[region_filter]['warehouses'])}
• Regional Description: {REGIONS[region_filter]['description']}"""
            
            success_msg = f"""Optimization Results with Regional Support, {distance_type} Distances & Weights Exported!

File: {file_path}

Distance Method: {distance_type} distances
{'🛣️ Using OpenRouteService for realistic road-based calculations' if self.engine.use_road_distances else '📏 Using geodesic (straight-line) calculations'}
{regional_info}

Scenarios Included ({scenario_count}):
{scenario_names}

Weight Information:
• Equipment weights included in all sheets
• Total weight calculations for each scenario
• Weight-based trip planning analysis
• Average total weight: {total_weight:.1f} lbs

Excel Sheets Created:
• Scenario Comparison - Side-by-side metrics with regional, weight & distance info
• Individual scenario sheets - Detailed pulls with regional, weight & distance info
  (Note: Sheet names are sanitized for Excel compatibility)
• Trip Comparison - Warehouse usage, regional distribution and weight analysis  
• Equipment Comparison - Item-by-item regional, weight & distance analysis

Equipment weights are stored in: {WEIGHTS_FILE}
OpenRouteService provides realistic road distances for better logistics planning.
Regional optimization helps optimize costs and comply with service agreements."""
            
            messagebox.showinfo("Export Complete", success_msg)
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not export: {e}")
            import traceback
            traceback.print_exc()

    # Settings Methods
    def populate_warehouse_tree(self):
        """Populate the warehouse coordinates tree"""
        for item in self.warehouse_tree.get_children():
            self.warehouse_tree.delete(item)
        
        for name, (lat, lon) in WAREHOUSE_COORDS.items():
            self.warehouse_tree.insert('', 'end', values=(name, f"{lat:.4f}", f"{lon:.4f}"))
    
    def verify_distances(self):
        """Verify distances from selected warehouse"""
        destination = self.dist_destination_var.get()
        if not destination:
            messagebox.showerror("Error", "Please select a destination warehouse")
            return
        
        try:
            self.distance_text.delete("1.0", tk.END)
            self.distance_text.insert(tk.END, "Calculating distances...\n")
            self.distance_text.update()
            
            distances = self.engine.verify_warehouse_distances(destination)
            
            self.distance_text.delete("1.0", tk.END)
            
            distance_type = "ROAD" if self.engine.use_road_distances else "GEODESIC"
            lines = [
                f"{distance_type} DISTANCE VERIFICATION FOR {destination}",
                f"Destination coordinates: {WAREHOUSE_COORDS[destination]}",
                f"Method: {'🛣️ OpenRouteService road distances' if self.engine.use_road_distances else '📏 Geodesic (straight line) distances'}",
                "-" * 60
            ]
            
            # Group warehouses by region for better display
            region_warehouses = {}
            for warehouse, coords in WAREHOUSE_COORDS.items():
                region = self.engine.get_warehouse_region(warehouse)
                if region not in region_warehouses:
                    region_warehouses[region] = []
                region_warehouses[region].append(warehouse)
            
            for region, warehouses in region_warehouses.items():
                lines.append(f"\n{region}:")
                for warehouse in sorted(warehouses):
                    if warehouse == destination:
                        lines.append(f"  {warehouse:<12}: {'0.0':>8} miles (LOCAL)")
                    else:
                        distance = distances.get(warehouse, 0)
                        if distance > 900000:
                            lines.append(f"  {warehouse:<12}: {'ERROR':>8}")
                        else:
                            distance_type_short = "ROAD" if self.engine.use_road_distances else "GEO"
                            lines.append(f"  {warehouse:<12}: {distance:>8.1f} miles ({distance_type_short})")
            
            lines.append("-" * 60)
            
            sorted_warehouses = sorted(distances.items(), key=lambda x: x[1])
            lines.append("WAREHOUSES BY DISTANCE:")
            for i, (wh, dist) in enumerate(sorted_warehouses, 1):
                region = self.engine.get_warehouse_region(wh)
                if dist == 0:
                    lines.append(f"{i:>2}. {wh:<12}: LOCAL ({region})")
                elif dist > 900000:
                    lines.append(f"{i:>2}. {wh:<12}: ERROR ({region})")
                else:
                    lines.append(f"{i:>2}. {wh:<12}: {dist:>8.1f} miles ({region})")
            
            lines.append("=" * 60)
            self.distance_text.insert(tk.END, "\n".join(lines))
            
        except Exception as e:
            messagebox.showerror("Error", f"Distance verification failed: {e}")


def main():
    """Main application entry point"""
    print("=" * 80)
    print("🚛 LOGISTICS MANAGEMENT SUITE v4.3 - With Regional Optimization & OpenRouteService")  
    print("=" * 80)
    print("🌍 NEW: Regional optimization with 2 predefined geographical regions")
    print("🛣️  Road-based distance calculation via OpenRouteService API")
    print("📏 Automatic fallback to geodesic distances if API unavailable")
    print("⚖️  Equipment weight management with JSON storage")
    print("📊 Enhanced optimization scenarios with regional constraints")
    print("=" * 80)
    
    root = tk.Tk()
    
    # Set application icon (if available)
    try:
        root.iconbitmap('logistics.ico')
    except:
        pass
    
    # Create and run application
    app = LogisticsManagementSuite(root)
    
    # Handle window closing
    def on_closing():
        app.save_settings()
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # Center window on screen
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")
    
    print("✅ Application initialized successfully!")
    print(f"🛣️  Road distances: {'ENABLED' if app.engine.use_road_distances else 'DISABLED (check API key)'}")
    print("🌍 Regional Features:")
    for region_name, region_data in REGIONS.items():
        print(f"   {region_name}: {len(region_data['warehouses'])} warehouses")
    print("📖 Instructions:")
    print("   1. Get free OpenRouteService API key at: https://openrouteservice.org/dev/#/signup")
    print("   2. Replace ORS_API_KEY in code with your actual API key")
    print("   3. Restart application to enable road-based distances")
    print("   4. Use Regional Filter dropdown to restrict pulls to specific regions")
    print("=" * 80)
    
    root.mainloop()


if __name__ == "__main__":
    main()